#import message_capturer
import sys, os, re
import json
import paho.mqtt.client as mqtt
import ssl
import threading
from openpyxl import load_workbook, Workbook
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QLabel, QPushButton, QAction, QMenu, QComboBox, QToolBar,
    QTableWidget, QSpacerItem, QTableWidgetItem, QDialog, QFrame, QFileDialog, QCheckBox, QSizePolicy, QHeaderView, QHBoxLayout, QSpinBox, QMessageBox, QDialogButtonBox
)
from PyQt5.QtGui import QPalette, QColor, QFont, QBrush, QIcon
from PyQt5.QtCore import Qt, pyqtSignal, QThread, QSize, QEvent, QObject, QRect
from openpyxl.styles import PatternFill
from datetime import datetime
from PyQt5.QtWidgets import QMessageBox, QInputDialog
import signal, time
import gui_config
from difflib import get_close_matches, SequenceMatcher

MAC_address_key = 'Mac Address'

def signal_handler(sig, frame):
    signal_names = {signal.SIGINT: "SIGINT (Ctrl+C)",
                    signal.SIGTERM: "SIGTERM (kill)"}
    print(f"{signal_names.get(sig, f'Signal {sig}')}, but ignoring to continue execution.")

# Register signal handlers
for sig in [signal.SIGINT, signal.SIGTERM]:
    signal.signal(sig, signal_handler)

class CheckableHeader(QHeaderView):
    def __init__(self, table_widget, parent=None):
        super().__init__(Qt.Horizontal, parent)
        self.table_widget = table_widget
        self.checkbox = QCheckBox(self)
        self.checkbox.setText("All")
        self.checkbox.setStyleSheet("margin-left: 2px; text-align: center;")
        self.checkbox.stateChanged.connect(self.on_state_changed)
        self.checked_rows = set()
        self.select_all_checkbox_triggered = False

    def resizeEvent(self, event):
        super().resizeEvent(event)
        section_width = self.sectionSize(0)
        header_height = self.height()
        checkbox_width = 150
        checkbox_height = 30
        x = int((section_width - checkbox_width) / 2)
        y = int((header_height - checkbox_height) / 2)

        self.checkbox.setGeometry(QRect(x, y, checkbox_width, checkbox_height))

    def on_state_changed(self, state):
        self.select_all_checkbox_triggered = True
        is_checked = (state == Qt.Checked)
        print("checked")
        table_window = self.parent().window()
        #self.checkbox.setText(f"{self.parent().window().selected_count} selected")
        print(f"table_window is {table_window}\n")
        if table_window:
            if(table_window.select_all_checkbox_triggered):
                table_window.select_all_states[table_window.current_page] = is_checked
                table_window.select_all_checkboxes(is_checked)
            print(state)
            if(self.select_all_checkbox_triggered) and is_checked is True:
                table_window.select_all_states[table_window.current_page] = is_checked
                table_window.select_all_checkboxes(is_checked)
                self.select_all_checkbox_triggered = False

            #self.checkbox.setText(f"{self.parent().window().selected_count} selected")

    def update_checkbox_count(self):
        print("enetered updating for header checkbox")
        table_window = self.parent().window()
        print(table_window.data)
        self.selected_count = sum(1 for row_data in table_window.data if row_data[5])
        print(f"count of selection is : {self.selected_count}")
        font = self.checkbox.font()
        font.setBold(self.selected_count > 0)
        self.checkbox.setFont(font)
        self.checkbox.setText(f"All ({self.selected_count} selected)")
        total_rows = len(table_window.data)
        print(total_rows)
        start_row = table_window.current_page * table_window.rows_per_page
        print(start_row)
        end_row = min(start_row + table_window.rows_per_page, total_rows)
        #print(f"length of self.data is {len(self.data)}")
        #end_row = (self.current_page + 1) * self.rows_per_page
        print(end_row)
        self.rows_in_a_page = end_row - start_row
        print(self.rows_in_a_page)
        print(f"current no of rows in page is {table_window.rows_in_a_page}")
        if self.selected_count < self.rows_in_a_page:
            #self.select_all_checkbox_triggered = False
            print(f"Selected count is {self.selected_count} and rows_in_page is {table_window.rows_in_a_page}")
            self.checkbox.setChecked(False)
        elif self.selected_count == self.rows_in_a_page:
            #self.select_all_checkbox_triggered = True
            print(f"Selected count is {self.selected_count} and rows_in_page is {table_window.rows_in_a_page}")
            self.checkbox.setChecked(True)
        return self.selected_count
    
    def set_row_checked(self, row, checked):
        if checked:
            print("added")
            self.checked_rows.add(row)
        else:
            print("discarded")
            self.checked_rows.discard(row)
        self.update_checkbox_count()


class DurationInputDialog(QDialog):
    def __init__(self, parent=None):
        super(DurationInputDialog, self).__init__(parent)
        self.setWindowTitle("Enter Duration")
        self.setFixedSize(400, 150)  # Adjust size to reduce extra space

        # Set layout with adjusted margins and spacing
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(20, 10, 20, 10)
        self.layout.setSpacing(5)

        # Create and style the label
        self.label = QLabel("Enter the duration (in seconds):")
        self.label.setAlignment(Qt.AlignLeft)  # Align text to the left
        font = QFont("Arial", 10)  # Set font style and size
        self.label.setFont(font)
        self.layout.addWidget(self.label)

        # Create and style the spin box
        self.spin_box = QSpinBox()
        self.spin_box.setRange(1, 9999)
        self.spin_box.setValue(1)
        self.spin_box.setFont(font)  # Set font style and size
        self.layout.addWidget(self.spin_box)

        # Create and style the OK button
        self.ok_button = QPushButton("OK")
        self.ok_button.setFont(font)  # Set font style and size
        self.ok_button.setFixedSize(60, 30)  # Adjust size of the button
        self.ok_button.clicked.connect(self.accept)
        self.layout.addWidget(self.ok_button)

        self.setLayout(self.layout)
        self.installEventFilter(self)

    def get_value(self):
        return self.spin_box.value()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Close:
            reply = QMessageBox.question(self, "Close", "Duration is not selected, Are you sure want to Exit?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                return super().eventFilter(obj, event)
            else:
                event.ignore()
                return True
        return super().eventFilter(obj, event)

class CountInputDialog(QDialog):
    def __init__(self, parent=None):
        super(CountInputDialog, self).__init__(parent)
        self.setWindowTitle("Enter Count")
        self.setFixedSize(400, 150)  # Adjust size to reduce extra space

        # Set layout with adjusted margins and spacing
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(20, 10, 20, 10)
        self.layout.setSpacing(10)

        # Create and style the label
        self.label = QLabel("Enter the count (number of times):")
        self.label.setAlignment(Qt.AlignLeft)  # Align text to the left
        font = QFont("Arial", 10)  # Set font style and size
        self.label.setFont(font)
        self.layout.addWidget(self.label)

        # Create and style the spin box
        self.spin_box = QSpinBox()
        self.spin_box.setRange(1, 9999)
        self.spin_box.setValue(1)
        self.spin_box.setFont(font)  # Set font style and size
        self.layout.addWidget(self.spin_box)

        # Create and style the OK button
        self.ok_button = QPushButton("OK")
        self.ok_button.setFont(font)  # Set font style and size
        self.ok_button.setFixedSize(60, 30)  # Adjust size of the button
        self.ok_button.clicked.connect(self.accept)
        self.layout.addWidget(self.ok_button)

        self.setLayout(self.layout)
        self.installEventFilter(self)

    def get_value(self):
        return self.spin_box.value()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Close:
            reply = QMessageBox.question(self, "Close", "Count is not selected, Are you sure want to exit?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                return super().eventFilter(obj, event)
            else:
                event.ignore()
                return True
        return super().eventFilter(obj, event)

'''class TopicSelectionDialog(QDialog):
    def __init__(self, parent = None):
        super(TopicSelectionDialog, self).__init__(parent)
        self.setWindowTitle("Settings")
        self.setFixedSize(300, 200)
        self.layout = QVBoxLayout()

        self.topic_label = QLabel("Select Topic:", self)
        self.layout.addWidget(self.topic_label)

        self.topic_combo = QComboBox()
        #self.topic_combo.addItems(["LED_GLOW", "LED_GLOW1"])
        self.topic_combo.addItems(gui_config.Topic_Names)
        self.topic_combo.setCurrentText(self.parent().window().selected_topic)
        self.topic_combo.setMinimumWidth(200)  # Increase the width of the combo box
        self.layout.addWidget(self.topic_combo)

        self.rows_label = QLabel("Rows per page:", self)
        self.layout.addWidget(self.rows_label)

        self.rows_spin_box = QSpinBox(self)
        self.rows_spin_box.setRange(1, 100)
        self.rows_spin_box.setValue(self.parent().window().rows_per_page)
        self.rows_spin_box.setMinimumWidth(100)  # Increase the width of the spin box
        self.layout.addWidget(self.rows_spin_box)'''

'''self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.layout.addWidget(self.ok_button)'''
'''self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.layout.addWidget(self.button_box)
        self.setLayout(self.layout)
        self.installEventFilter(self)
        

    def get_selected_topic(self):
        return self.topic_combo.currentText()
    
    def get_rows_per_page(self):
        return self.rows_spin_box.value()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Close:
            reply = QMessageBox.question(self, "Close", "The changes you made may get lost, Are you sure want to exit?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                return super().eventFilter(obj, event)
            else:
                event.ignore()
                return True
        return super().eventFilter(obj, event)'''

class TopicSelectionDialog(QDialog):
    def __init__(self, parent=None):
        super(TopicSelectionDialog, self).__init__(parent)
        self.setWindowTitle(" ")
        self.setFixedSize(600, 300)
        self.setStyleSheet("background-color: lightgrey;")
 
        # Custom Title Bar
        title_bar = QWidget(self)
        title_layout = QHBoxLayout()
        title_label = QLabel("Settings", self)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        title_layout.addWidget(title_label)
 
        title_layout.setContentsMargins(0, 5, 0, 10)  # Reduced top margin
        title_bar.setLayout(title_layout)
 
        self.layout = QVBoxLayout()
        self.layout.addWidget(title_bar)
 
        # Page Parameters
        self.rows_layout = QHBoxLayout()
        self.rows_label = QLabel("Page Size:", self)
        self.rows_label.setStyleSheet("font-size: 18px;")
        self.rows_layout.addWidget(self.rows_label)
 
        self.rows_spin_box = QSpinBox(self)
        self.rows_spin_box.setRange(1, 100)
        self.rows_spin_box.setValue(self.parent().window().rows_per_page)  # Default value
        self.rows_spin_box.setMinimumWidth(120)
        self.rows_spin_box.setAlignment(Qt.AlignCenter)  # Center-align the text
        self.rows_spin_box.setStyleSheet("height: 30px; border: 2px solid black; font-size: 18px;")
        self.rows_layout.addWidget(self.rows_spin_box)
        self.layout.addLayout(self.rows_layout)
 
        # Divider Line
        self.divider_line = QFrame(self)
        self.divider_line.setFrameShape(QFrame.HLine)
        self.divider_line.setFrameShadow(QFrame.Sunken)
        self.divider_line.setStyleSheet("color: grey;")
        self.divider_line.setFixedHeight(2)
        self.layout.addWidget(self.divider_line)
 
        # LED Glow Test Settings Title
        self.test_settings_label = QLabel("LED Glow Test Settings", self)
        self.test_settings_label.setAlignment(Qt.AlignCenter)
        self.test_settings_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        self.layout.addWidget(self.test_settings_label)
 
        # Run Test Parameters
        self.topic_layout = QHBoxLayout()
        self.topic_label = QLabel("Topic Name:", self)
        self.topic_label.setStyleSheet("font-size: 18px;")
        self.topic_layout.addWidget(self.topic_label)
 
        self.topic_combo = QComboBox()
        self.topic_combo.setEditable(True)
        self.topic_combo.lineEdit().setAlignment(Qt.AlignCenter)
        self.topic_combo.addItems(gui_config.Topic_Names)  # Example topics
        self.topic_combo.setCurrentText(self.parent().window().selected_topic)
        self.topic_combo.setMinimumWidth(200)  # Increased width
        #self.topic_combo.setAlignment(Qt.AlignCenter)
        self.topic_combo.setStyleSheet("height: 30px; border: 2px solid black; font-size: 18px; text-align: center;")
        self.topic_layout.addWidget(self.topic_combo)
        self.layout.addLayout(self.topic_layout)
 
        # Blink Duration
        self.blink_duration_layout = QHBoxLayout()
        self.blink_duration_label = QLabel("LED Blink Duration (seconds):", self)
        self.blink_duration_label.setStyleSheet("font-size: 18px;")
        self.blink_duration_layout.addWidget(self.blink_duration_label)
 
        self.blink_duration_spin_box = QSpinBox(self)
        self.blink_duration_spin_box.setRange(1, 60)
        self.blink_duration_spin_box.setValue(self.parent().window().duration)  # Default value
        self.blink_duration_spin_box.setMinimumWidth(120)
        self.blink_duration_spin_box.setAlignment(Qt.AlignCenter)
        self.blink_duration_spin_box.setStyleSheet("height: 30px; border: 2px solid black; font-size: 18px;")
        self.blink_duration_layout.addWidget(self.blink_duration_spin_box)
        self.layout.addLayout(self.blink_duration_layout)
 
        # Blink Count
        self.blink_count_layout = QHBoxLayout()
        self.blink_count_label = QLabel("LED Blink Count:", self)
        self.blink_count_label.setStyleSheet("font-size: 18px;")
        self.blink_count_layout.addWidget(self.blink_count_label)
 
        self.blink_count_spin_box = QSpinBox(self)
        self.blink_count_spin_box.setRange(1, 100)
        self.blink_count_spin_box.setValue(self.parent().window().count)  # Default value
        self.blink_count_spin_box.setMinimumWidth(120)
        self.blink_count_spin_box.setAlignment(Qt.AlignCenter)
        self.blink_count_spin_box.setStyleSheet("height: 30px; border: 2px solid black; font-size: 18px;")
        self.blink_count_layout.addWidget(self.blink_count_spin_box)
        self.layout.addLayout(self.blink_count_layout)
 
        # Adding spacing between labels and spin boxes
        for layout in [self.rows_layout, self.topic_layout, self.blink_duration_layout, self.blink_count_layout]:
            layout.setSpacing(30)  # Increased space between labels and spin boxes
 
        # Button Box
        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.button_box.setCenterButtons(True)  # Center the buttons
        self.layout.addWidget(self.button_box)
 
        self.setLayout(self.layout)
        self.installEventFilter(self)
 
    def get_selected_topic(self):
        return self.topic_combo.currentText()
 
    def get_rows_per_page(self):
        return self.rows_spin_box.value()
   
    def get_blink_duration(self):
        return self.blink_duration_spin_box.value()
 
    def get_blink_count(self):
        return self.blink_count_spin_box.value()
 
    def eventFilter(self, obj, event):
        if event.type() == QEvent.Close:
            reply = QMessageBox.question(self, "Close",
                "The changes you made may get lost. Are you sure you want to exit?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                return super().eventFilter(obj, event)
            else:
                event.ignore()
                return True
        return super().eventFilter(obj, event)



class ErrorDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Error')
        self.message_label = QLabel()
        layout = QVBoxLayout()
        layout.addWidget(self.message_label)
        self.setLayout(layout)


    def set_message(self, message):
        self.message_label.setText(message)


class MqttHandler(QThread):
    mqtt_message_signal = pyqtSignal(str, str)


    def __init__(self, broker, port, username, password):
        super().__init__()
        self.broker = broker
        self.port = port
        self.username = username
        self.password = password
        self.client = mqtt.Client(client_id="unique_client_id")
        self.client.username_pw_set(self.username, self.password)
        self.client.tls_set(cert_reqs=ssl.CERT_NONE)
        self.client.tls_insecure_set(True)
        self.client.on_connect = self.on_connect
        self.client.on_message = self.on_message


    def run(self):
        self.client.connect(self.broker, self.port, 60)
        self.client.loop_forever()


    def on_connect(self, client, userdata, flags, rc):
        if rc == 0:
            print("Connected to MQTT Broker!")
            client.subscribe("#", qos=1)
        else:
            print(f"Failed to connect, return code {rc}")


    '''def on_message(self, client, userdata, msg):
        topic = msg.topic
        payload = msg.payload.decode()
        self.mqtt_message_signal.emit(topic, payload)'''
    def on_message(self, client, userdata, msg):
        topic = msg.topic
        try:
            payload = msg.payload.decode('utf-8') # Try decoding with UTF-8
        except UnicodeDecodeError:
            # Handle non-UTF-8 payloads or binary data
            payload = f"Binary data received: {msg.payload}"
    
        self.mqtt_message_signal.emit(topic, payload)


class DarkWindow(QMainWindow):
    def __init__(self):
        super().__init__()


        self.setWindowTitle('Device Tester')
        self.setGeometry(100, 100, 1200, 800)


        palette = QPalette()
        base_color = QColor(35, 35, 35)
        alternate_color = QColor(53, 53, 53)
        palette.setColor(QPalette.Window, QColor(53, 53, 53))
        palette.setColor(QPalette.WindowText, QColor(255, 255, 255))
        palette.setColor(QPalette.Base, QColor(35, 35, 35))
        palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
        palette.setColor(QPalette.ToolTipBase, QColor(255, 255, 255))
        palette.setColor(QPalette.ToolTipText, QColor(255, 255, 255))
        palette.setColor(QPalette.Text, QColor(255, 255, 255))
        palette.setColor(QPalette.Button, QColor(53, 53, 53))
        palette.setColor(QPalette.ButtonText, QColor(255, 255, 255))
        palette.setColor(QPalette.BrightText, QColor(255, 0, 0))
        palette.setColor(QPalette.Link, QColor(42, 130, 218))
        palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        palette.setColor(QPalette.HighlightedText, QColor(0, 0, 0))


        self.setPalette(palette)

        self.setMinimumSize(400, 200)
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)
        self.current_page = 0
        self.rows_per_page = gui_config.rows_in_page
        self.total_rows = 0
        self.select_all_states = {}
        self.prev_pages = 1
        self.rows_in_a_page = 0
        self.duration = gui_config.blink_duration
        self.count = gui_config.blink_count

        self.select_all_checkbox_triggered = False
        self.layout = QVBoxLayout(central_widget)

        '''self.upload_button = QPushButton('Upload File')
        self.upload_button.setStyleSheet("background-color: grey; color: white; font-size:12px;")
        self.upload_button.clicked.connect(self.upload_file)
        #layout.addWidget(self.upload_button)'''
        self.upload_done = False

        self.file_label = QLabel('')
        self.file_label.setStyleSheet("color: white;")
        self.layout.addWidget(self.file_label)


        self.table_widget = QTableWidget(self)
        self.table_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.layout.addWidget(self.table_widget)
        self.table_widget.setVisible(False)
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setStyleSheet("""
            QTableWidget {
                alternate-background-color: #444444;
                background-color: #2B2B2B;
            }
            QHeaderView: {
                background-color: #353535;
                color: white;
            }
        """)

        #self.header = CheckableHeader(self.table_widget, self)
        #self.table_widget.setHorizontalHeader(self.header)
        self.checkable_header = CheckableHeader(self.table_widget)  # Keep a reference
        self.table_widget.setHorizontalHeader(self.checkable_header)
        #self.checkbox_count_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        self.status_label = QLabel()
        self.status_label.setStyleSheet("color: white;")
        self.layout.addWidget(self.status_label)

        '''button_layout = QHBoxLayout()

        self.download_button = QPushButton('Download Excel')
        self.download_button.setStyleSheet("background-color: grey; color: white; font-size:12px;")
        self.download_button.clicked.connect(self.download_excel)
        self.download_button.setVisible(False)
        #button_layout.addWidget(self.download_button)


        self.run_test_button = QPushButton('Run Test')
        self.run_test_button.setStyleSheet("background-color: grey; color: white; font-size:12px;")
        self.run_test_button.clicked.connect(self.run_test)
        self.run_test_button.setVisible(False)
        #button_layout.addWidget(self.run_test_button)

        button_layout.setStretch(0,1)
        button_layout.setStretch(1,1)
        self.download_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.run_test_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        layout.addLayout(button_layout)'''

        central_widget.setLayout(self.layout)
        self.toolbar = QToolBar()
        self.toolbar.setStyleSheet("""
            QToolBar {
                border: 2px solid white; /* Set the same background color for the toolbar */
                padding: 0px;
                spacing: 3px;
                height: 70px;
            }
            QToolButton {
                background-color: transparent;  /* Ensure buttons have the same background color */
                color: white;
                font-size: 12px;         /* Increase the font size */
                margin: 0px;
                padding: 2px;            /* Add some padding to avoid cramped text */
            }
            QToolButton::icon {
            padding-top: 5px;
            }
        """)
        self.toolbar.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)
        self.toolbar.setMovable(False)
        self.toolbar.setFloatable(False)
        #spacer = QWidget()
        #spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        #self.toolbar.addWidget(spacer)
        #self.toolbar.setStyleSheet("background-color: #E0E0E0 ;;")  # You can adjust the color to your preference
        self.addToolBar(Qt.TopToolBarArea, self.toolbar)
        self.toolbar.setIconSize(QSize(48,48))

        self.upload_action = QAction(QIcon('upload3.png'), 'Upload', self)
        self.upload_action.triggered.connect(self.upload_file)
        self.run_test_action = QAction(QIcon('run.png'), 'Run Test', self)
        self.run_test_action.triggered.connect(self.run_test)
        self.download_action = QAction(QIcon('download.png'), 'Download', self)
        self.download_action.triggered.connect(self.download_excel)
        self.settings_action = QAction(QIcon('images.png'), 'Settings', self)
        self.settings_action.triggered.connect(self.open_settings_dialog)
        self.toolbar.addAction(self.upload_action)
        self.toolbar.addAction(self.run_test_action)
        self.toolbar.addAction(self.download_action)
        self.toolbar.addAction(self.settings_action)

        self.nav_layout = QHBoxLayout()
        self.nav_layout.setContentsMargins(0, 0, 0, 0)  # No margins around the layout
        self.nav_layout.setSpacing(10)  # No spacing between items

        # Add stretchable space before the previous button
        self.nav_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))

        self.page_info_label = QLabel(self)
        self.page_info_label.setAlignment(Qt.AlignCenter)
        self.page_info_label.setStyleSheet("color: white; font-size: 14px;")  # Padding for spacing

        self.first_button = QPushButton(self)
        self.first_button.setIcon(QIcon('first_page.png'))
        self.first_button.clicked.connect(self.go_to_first_page)
        self.first_button.setVisible(False)
        self.first_button.setStyleSheet("border: none; background: none;")

        self.prev_button = QPushButton(self)
        self.prev_button.setIcon(QIcon('left.png'))
        self.prev_button.clicked.connect(self.load_previous_page)
        self.prev_button.setVisible(False)
        self.prev_button.setStyleSheet("border: none; background: none;")  # Transparent background


        self.next_button = QPushButton(self)
        self.next_button.setIcon(QIcon('right.png'))        
        self.next_button.clicked.connect(self.load_next_page)
        self.next_button.setVisible(False)
        self.next_button.setStyleSheet("border: none; background: none;")  # Transparent background

        self.last_button = QPushButton(self)
        self.last_button.setIcon(QIcon('last_page.png'))
        self.last_button.clicked.connect(self.go_to_last_page)
        self.last_button.setVisible(False)
        self.last_button.setStyleSheet("border: none; background: none;")

        self.first_button.setFixedSize(30, 30)
        self.prev_button.setFixedSize(30, 30)  # Set a fixed size for the left icon
        self.next_button.setFixedSize(30, 30)
        self.last_button.setFixedSize(30, 30)

        #self.prev_button.setContentsMargins(100, 0, 1000, 0)  # Small space to the right
        #self.page_info_label.setContentsMargins(100, 0, 100, 0)  # Small space on both sides
        #self.next_button.setContentsMargins(100, 0, 0, 0)  # Small space to the left




        self.nav_layout.addWidget(self.first_button, alignment=Qt.AlignCenter)
        self.nav_layout.addWidget(self.prev_button, alignment=Qt.AlignCenter)
        self.nav_layout.addWidget(self.page_info_label, alignment=Qt.AlignCenter)
        self.nav_layout.addWidget(self.next_button, alignment=Qt.AlignCenter)
        self.nav_layout.addWidget(self.last_button, alignment=Qt.AlignCenter)
        
        #if self.upload_done is True:
        
        #self.nav_layout.setContentsMargins(50, 0, 0, 0)  # Remove all margins
        #self.nav_layout.setSpacing(100)
        # Add stretchable space after the next button
        self.nav_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
        self.layout.addLayout(self.nav_layout)

        self.selected_topic = 'LED_GLOW'

        self.installEventFilter(self) 
        self.data = []

        self.MQTT_BROKER = "103.162.246.109"
        #self.MQTT_BROKER = "192.168.100.18"
        self.MQTT_PORT = 8883
        self.MQTT_USERNAME = "mqtt"
        self.MQTT_PASSWORD = "mqtt2022"

        self.mqtt_handler = MqttHandler(self.MQTT_BROKER, self.MQTT_PORT, self.MQTT_USERNAME, self.MQTT_PASSWORD)
        #self.boot_up_logic()
        self.mqtt_handler.mqtt_message_signal.connect(self.handle_mqtt_message)
        self.mqtt_handler.start()


        self.selected_count = 0
        print(f"selected devices count = {self.selected_topic}")
        self.topic_dialog = TopicSelectionDialog(self)
        #self.checkbox_count_label = QLabel("Selected: 0")
        #layout.addWidget(self.checkbox_count_label)
        #self.checkbox_count_label.setVisible(False)
        #self.boot_up_logic()

    #def boot_up_logic(self):
        #self.message = message_capturer.ssh_connect_and_execute(self.server, self.port, self.username, self.password, self.command, self.path)
        #print(self.message)

    def update_page_info(self):
        total_pages = (self.total_rows + self.rows_per_page - 1) // self.rows_per_page
        if total_pages > 1:
            self.first_button.setVisible(True)
            self.prev_button.setVisible(True)
            self.next_button.setVisible(True)
            self.last_button.setVisible(True)
            print("Added first and last button")
            self.page_info_label.setText(f"Page {self.current_page + 1} of {total_pages}")
            self.first_button.setEnabled(self.current_page > 0)
            self.prev_button.setEnabled(self.current_page > 0)
            self.next_button.setEnabled(self.current_page < total_pages - 1)
            self.last_button.setEnabled(self.current_page < total_pages - 1)
        else:
            self.page_info_label.setText('')
            #if self.nav_layout.isVisible():
                #self.clear_layout(self.nav_layout)
                #self.nav_layout.setVisible(False)
            # Always keep buttons visible but change their appearance
            self.first_button.setVisible(False)
            self.prev_button.setVisible(False)
            self.next_button.setVisible(False)
            self.last_button.setVisible(False)
        
        # Optional: Set a different style for disabled buttons
        if not self.first_button.isEnabled():
            self.first_button.setStyleSheet("opacity: 0.5;")
        else:
            self.first_button.setStyleSheet("")

        if not self.prev_button.isEnabled():
            self.prev_button.setStyleSheet("opacity: 0.5;")  # Make it look disabled
        else:
            self.prev_button.setStyleSheet("")  # Reset style

        if not self.next_button.isEnabled():
            self.next_button.setStyleSheet("opacity: 0.5;")  # Make it look disabled
        else:
            self.next_button.setStyleSheet("")  # Reset style

        if not self.last_button.isEnabled():
            self.last_button.setStyleSheet("opacity: 0.5;")
        else:
            self.last_button.setStyleSheet("")
        

    def go_to_first_page(self):
        self.current_page = 0
        self.update_table()
        self.update_page_info()
        #self.load_page_data()

    def go_to_last_page(self):
        total_page = (self.total_rows + self.rows_per_page - 1) // self.rows_per_page
        self.current_page = total_page - 1
        self.update_table()
        self.update_page_info()
        #self.load_page_data()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Close:
            reply = QMessageBox.question(self, "Exit", "Exit GUI?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                return super().eventFilter(obj, event)
            else:
                event.ignore()
                return True
        return super().eventFilter(obj, event)
    
    def handle_mqtt_message(self, topic, payload):
        try:
            print(f"Topic = {topic}")
            print(f"receieved payload is :{payload}")
            message = json.loads(payload)
            dev_id = message.get('devID')
            print(f"dev_id is {dev_id}")
            if dev_id:
                #mac_address = ':'.join([dev_id[i:i + 2] for i in range(0, len(dev_id), 2)])
                mac_address = dev_id
                self.update_table_with_mqtt_data(mac_address, message.get('data'))
            else:
                print("Received message, but no devID found in payload")

            print(f"mac_address = {mac_address}, data = {message.get('data')}")
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON: {e}")
        QApplication.processEvents()
        

    def open_settings_dialog(self):
        dialog = TopicSelectionDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            self.selected_topic = dialog.get_selected_topic()
            self.rows_per_page = dialog.get_rows_per_page()
            self.duration = dialog.get_blink_duration()
            self.count = dialog.get_blink_count()
            if(self.upload_done):
                self.status_label.setText(f"Selected Topic for LED Test: {self.selected_topic}")
                self.update_page_info()
                self.update_table()
                print(f"the total rows are {self.total_rows}")
                total_pages = (self.total_rows + self.rows_per_page - 1) // self.rows_per_page
                print(f"Total_pages before are {self.prev_pages} and now it is {total_pages}") 
                if(total_pages < self.prev_pages):
                    self.go_to_last_page() 
                if(total_pages == 0):
                    QMessageBox.warning(self, "No Valid Addresses", "No valid addresses are found in the given Excel sheet.")
                self.prev_pages = total_pages
                

    '''def run_test(self):
        selected_rows = self.get_selected_rows()
        if not selected_rows:
            error_dialog = ErrorDialog(self)
            error_dialog.set_message("Please select at least one row.")
            error_dialog.exec_()
            return

        # Use the custom DurationInputDialog
        duration_dialog = DurationInputDialog(self)
        if duration_dialog.exec_() == QDialog.Accepted:
            duration =.get_value()
        else:
            return

        # Use the custom CountInputDialog
        count_dialog = CountInputDialog(self)
        if count_dialog.exec_() == QDialog.Accepted:
            count = count_dialog.get_value()
        else:
            return
        
        duration = TopicSelectionDialog.get_blink_duration()
        count = TopicSelectionDialog.get_blink_count()

        for row_index in selected_rows:
            mac_address = self.data[row_index][0]
            mac_address_without_colon = ''.join(mac_address.split(':'))
            payload = f"LG {duration},{count}"
            self.publish_message(mac_address, payload)
        QApplication.processEvents()'''
    
    def run_test(self):
        selected_rows = self.get_selected_rows()
        if not selected_rows:
            error_dialog = ErrorDialog(self)
            error_dialog.set_message("Please select at least one row.")
            error_dialog.exec_()
            return
 
        # Open Topic Selection Dialog to get parameters
        #topic_dialog = TopicSelectionDialog(self)
        #if topic_dialog.exec_() != QDialog.Accepted:
            #return  # User canceled the dialog
 
        # Retrieve values from the dialog
        #self.selected_topic = self.topic_dialog.get_selected_topic()
        #self.duration = self.topic_dialog.get_blink_duration()
        #self.count = self.topic_dialog.get_blink_count()
 
        # Constructing the message
        '''message = (
            f"<b>Publishing LED glow test, '{self.selected_topic}' for the selected {self.selected_count} devices.</b><br>\n"
            f"<b>Blink duration: {self.duration} seconds, Blink count: {self.count}</b><br>\n"
            "Note: Go to settings to change any of the parameters."
        )'''
        self.selected_count = sum(1 for row_data in self.data if row_data[5])


        message = (
        f"<div style='text-align: center;'>"
        f"<b>Publishing LED glow test, '{self.selected_topic}' for the selected {self.selected_count} devices.</b><br>"
        f"<b>Blink duration: {self.duration} seconds, Blink count: {self.count}</b><br>"
        "Note: Go to settings to change any of the parameters."
        "</div>"
    )
 
        # Create a custom dialog for confirmation
        custom_dialog = QDialog(self)
        custom_dialog.setWindowTitle("Run Test")
        custom_dialog.setFixedSize(550, 150)  # Adjust size as necessary
 
        layout = QVBoxLayout()
   
        message_label = QLabel(message)
        message_label.setWordWrap(True)
        message_label.setStyleSheet("font-size: 15px")
        layout.addWidget(message_label)
 
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, custom_dialog)
        button_box.setCenterButtons(True)
        layout.addWidget(button_box)
 
        custom_dialog.setLayout(layout)
 
        # Connect buttons
        button_box.accepted.connect(custom_dialog.accept)
        button_box.rejected.connect(custom_dialog.reject)
 
        # Show the custom dialog
        reply = custom_dialog.exec_()
 
        print(f"selected devices count = {self.selected_count}")

        if reply == QDialog.Accepted:
            for row_index in selected_rows:
                mac_address = self.data[row_index][0]
                mac_address_without_colon = ''.join(mac_address.split(':'))
                payload = f"LG {self.duration},{self.count}"
                self.publish_message(mac_address, payload)
            QApplication.processEvents()
        else:
            confirmation_reply = QMessageBox.question(self, "Confirm Exit", "Are you sure you want to quit running the test?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            #return  # User chose to cancel

            if confirmation_reply == QMessageBox.Yes:
                return 

        

    def get_selected_rows(self):
        selected_rows = []
        for row_index, row_data in enumerate(self.data):
            if row_data[5]:
                #print(row_data[5])
                selected_rows.append(row_index)
        return selected_rows
    
    def publish_message(self, mac_address, payload):
        topic = self.selected_topic
        message = {
            "devID": mac_address.upper(),
            "data": payload
        }
        self.mqtt_handler.client.publish(topic, json.dumps(message), qos=1)
        QApplication.processEvents()
        #print(f"Published message '{message}' to topic '{topic}'")
 
        
    def update_table_with_mqtt_data(self, mac_address, data):
        data_found = False
        mac_address_lower = mac_address.lower()
    
        # Check if 'LG' appears in the data
        is_LG_data = any(part.strip().lower().startswith("lg") for part in data.split(','))
        print(f"LG data is {is_LG_data}")

        for row_index, row_data in enumerate(self.data):
            if row_data[0] is not None and row_data[0].lower() == mac_address_lower:
                data_found = True
                self.update_checkbox(row_index, data)
                print(data)

                # Only move the row if the data does not contain "LG"
                if any(row_data[1:]):
                    print("entering to pop")
                    row = self.data.pop(row_index)
                    self.data.insert(0, row)
                break

        if not data_found:
            print(data)
            new_row = [mac_address, False, False, False, False, False]
        
            # Add the new row to the end if data is LG
            if is_LG_data:
                self.data.append(new_row)  # Append to the end for LG data
            else:
                print(f"before popping no of rows will be {len(self.data)}")
                last_row = self.data.pop(len(self.data)-1)
                print(f"After popping the last row, no of rows will be {len(self.data)}")
                print(f"Last row is {last_row}")
                print("pushed to first row")
                self.data.insert(0, new_row)  # Push new row to the top
                print(f"before adding the popped row to last, no of rows will be {len(self.data)}")
                print("adding row to last")
                self.data.append(last_row)
                self.total_rows = len(self.data)
                print(f"after adding the popped row to last no of rows will be {len(self.data)}")
                #self.update_page_info()
            self.update_page_info()
            self.update_checkbox(len(self.data)-1 if is_LG_data else 0, data)

        self.update_table()



    def update_checkbox(self, row_index, data):
        if data == 100:
            self.data[row_index][1] = True
            move_row = True
        elif data == "BootUp":
            self.data[row_index][3] = True
            move_row = True
        elif data == "SW":
            self.data[row_index][2] = True
            move_row = True
        #if move_row == True:
         #   print(self.data[row_index])
        #elif isinstance(data, (int, float)) and 0<= data <= 100:
        #    self.data[row_index][3] = True


    def upload_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx)", options=options)
        if file_name:
            self.load_data(file_name)
            self.status_label.setText('Excel sheet is uploading. Please wait...')
            self.file_label.setText(f'Selected File: {file_name}')
            #self.header = CheckableHeader(self.table_widget, self)
            #self.table_widget.setHorizontalHeader(self.header)
            font = self.table_widget.horizontalHeader().checkbox.font()
            font.setBold(False)
            self.table_widget.horizontalHeader().checkbox.setFont(font)
            self.table_widget.horizontalHeader().checkbox.setText(f"All")
            self.select_all_states = {}
            #self.checkbox_count_label.setText(f"No MAC Addresses are selected")
            self.update_table()
            self.table_widget.setVisible(True)
            self.download_action.setVisible(True)
            self.run_test_action.setVisible(True)
            total_pages = (self.total_rows + self.rows_per_page - 1) // self.rows_per_page
            print(f"Previous total pages are {self.prev_pages} and current total pages are {total_pages}")
            if(total_pages == 1):
                self.go_to_first_page()
            elif(total_pages > self.prev_pages):
                self.first_button.setVisible(True)
                self.prev_button.setVisible(True)
                self.next_button.setVisible(True)
                self.last_button.setVisible(True)
            elif(total_pages == 0):
                QMessageBox.warning(self, "No Valid Addresses", "No valid addresses are found in the given Excel sheet.")
            elif(total_pages<self.prev_pages):
                self.go_to_last_page()
            else:
                self.go_to_first_page()
            
            self.status_label.setText('Upload done.')
            #self.setLayout(self.nav_layout)
            #self.table_widget.setVisible(True)
            self.upload_done = True
            self.prev_pages = total_pages
            #self.checkbox_count_label.setVisible(True)

    def find_mac_address_column(self, sheet):
        possible_headers = {'mac address', 'mac-address', 'mac_address'}
        headers = [cell.value.strip().lower() for cell in sheet[1] if cell.value]
    
        print(f"Headers found: {headers}")
 
        best_match_idx = None
        highest_ratio = 0.0
 
        for idx, header in enumerate(headers):
            print(f"Checking header: {header}")
            match = get_close_matches(header, possible_headers, n=1, cutoff=0.8)
            print(f"Match found: {match}")
        
            if match:
                ratio = SequenceMatcher(None, header, match[0]).ratio()
                print(f"Similarity ratio: {ratio}")
                if ratio > highest_ratio:
                    highest_ratio = ratio
                    best_match_idx = idx
 
        if best_match_idx is not None:
            print(f"Best match column index: {best_match_idx}")
        else:
            print("No match found.")
 
        return best_match_idx
    '''def load_data(self, file_name):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active

        self.data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            mac_address = row[1]
            #print(f"Read MAC Address: {mac_address}")S
            if mac_address and self.is_valid_mac_address(mac_address):
                self.data.append([mac_address, False, False, False, False, False])'''

    def load_data(self, file_name):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active

        mac_addresses = set()
        duplicates = set()
        complete_mac_address = set()
        mac_address_column = self.find_mac_address_column(sheet)
        print(mac_address_column)
        self.data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            mac_address = row[mac_address_column]
            if mac_address:
                normalized_mac = mac_address.strip().lower()
                complete_mac_address.add(normalized_mac)
                if normalized_mac in mac_addresses:
                    duplicates.add(normalized_mac)
                else:
                    mac_addresses.add(normalized_mac)

    # Print unique MAC addresses
        print("Unique MAC addresses:")
        for mac in complete_mac_address:
            print(mac)

        if duplicates:
            reply = QMessageBox.question(
                self, 'Duplicate MAC Addresses',
                f"Duplicate MAC addresses found: {', '.join(duplicates)}.\n\nDo you want to remove the duplicates?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                self.remove_duplicates_load(complete_mac_address)
                #self.data = [[mac.strip().upper(), False, False, False, False, False] for mac in complete_mac_address]
            else:
                self._load_data_from_file(file_name, mac_address_column, remove_invalid=True)
        else:
            self._load_data_from_file(file_name, mac_address_column, remove_invalid=True)
        self.total_rows = len(self.data)
    
    def remove_duplicates_load(self, mac_address_all):
        self.data = []
        for mac in mac_address_all:
            if self.is_valid_mac_address(mac.strip().upper()):
                    self.data.append([mac.strip().upper(), False, False, False, False, False])
            
        #for mac_address in mac_addresses:
         #   print(mac_address)
          #  if mac_address and self.is_valid_mac_address(mac_address):
           #     self.data.append([mac_address, False, False, False, False, False])
 
    '''def _load_data_from_file(self, file_name, mac_address_column):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active
        self.data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            mac_address = row[mac_address_column]
            if mac_address and self.is_valid_mac_address(mac_address):
                self.data.append([mac_address, False, False, False, False, False])'''
    def _load_data_from_file(self, file_name, mac_address_column, remove_invalid=True):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active
        self.data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            mac_address = row[mac_address_column]
            if mac_address:
                normalized_mac = mac_address.strip().lower()
                if remove_invalid:
                    if self.is_valid_mac_address(normalized_mac.upper()):
                        print(f"{normalized_mac} is valid")
                        self.data.append([normalized_mac.upper(), False, False, False, False, False])
                    #else:
                    #    print(f"{normalized_mac} is Invalid")
                    #    self.data.append([normalized_mac.upper(), False, False, False, False, False])


    def is_valid_mac_address(self, mac_address):
        print(f"{mac_address} is checking for validity")
        if not re.match(r'^([0-9A-Fa-f]{2}:){5}[0-9A-Fa-f]{2}$', mac_address):
            return False
        if not mac_address.startswith("AB:05:03"):
            return False
        return True
        #stripped_mac = re.sub(r'[^a-fA-F0-9]', '', mac_address)
        #return len(stripped_mac) == 12 and all(c in '0123456789abcdefABCDEF' for c in stripped_mac)

    
    
    def update_table(self):
        self.table_widget.clear() # Clears all items from the table widget
        self.table_widget.setRowCount(len(self.data)) # Sets the number of rows in the table to the length of self.data
        self.table_widget.setColumnCount(6) # Sets the number of columns in the table to the length of self.data
        self.table_widget.setAlternatingRowColors(True) # Enables alternating row colors for the table
        self.table_widget.setHorizontalHeaderLabels(['', 'MAC Address', 'Battery Status', 'Switch Press', 'Bootup Test', 'LED Glow']) # Sets the horizontal header labels for the columns
        
        # Configure the horizontal header to resize sections to fit their contents
        start_row = self.current_page * self.rows_per_page
        #print(f"length of self.data is {len(self.data)}")
        end_row = (self.current_page + 1) * self.rows_per_page
        #print(f"start_row is {start_row}")
        #print(f"end_row is {end_row}")
        page_data = self.data[start_row:end_row]
        select_all_checked = self.select_all_states.get(self.current_page, False)
        self.table_widget.setRowCount(len(page_data))

        self.checkable_header.checkbox.setChecked(select_all_checked)
        self.table_widget.horizontalHeader().setVisible(True)

        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        #print(page_data)

        #selected_count = 0
        
        # Iterate over each row in self.data and poplate the table
        for row_index, row_data in enumerate(page_data):
            if len(row_data) >= 6:

                # Creates and configure a checkbox for the first column
                checkbox= QCheckBox()
                checkbox.setChecked(row_data[5])
                checkbox.stateChanged.connect(lambda state, idx=row_index + start_row: self.handle_checkbox_change(idx, state))
                self.center_checkbox_in_cell(row_index, 0, checkbox)

                # Creates and configure an item for the MAC Address Column
                #print(row_data[1])
                mac_item = QTableWidgetItem(row_data[0])
                font = mac_item.font()
                font.setPointSize(12)
                mac_item.setFont(font)
                mac_item.setForeground(QBrush(QColor(255, 255, 255)))
                mac_item.setTextAlignment(Qt.AlignCenter)
                mac_item.setFlags(mac_item.flags() & ~Qt.ItemIsEditable | Qt.ItemIsSelectable)
                self.table_widget.setItem(row_index, 1, mac_item)

                # Created and configure a checkbox for the Battery Status column
                battery_status = QCheckBox()
                battery_status.setChecked(row_data[1])
                battery_status.setEnabled(False)
                battery_status.setStyleSheet("""
                    QCheckBox::indicator {
                        width: 15px;
                        height: 15px;
                    }
                    QCheckBox::indicator::unchecked {
                        background-color: white;
                        border: 1px solid white;
                    }
                    QCheckBox::indicator:checked {
                        background-color: green;
                    }
                """)
                self.center_checkbox_in_cell(row_index, 2, battery_status)


                switch_press = QCheckBox()
                switch_press.setChecked(row_data[2])
                switch_press.setEnabled(False)
                switch_press.setStyleSheet("""
                    QCheckBox::indicator {
                        width: 15px;
                        height: 15px;
                    }
                    QCheckBox::indicator::unchecked {
                        background-color: white;
                        border: 1px solid white;
                    }
                    QCheckBox::indicator:checked {
                        background-color: green;
                    }
                """)
                self.center_checkbox_in_cell(row_index, 3, switch_press)


                bootup_test = QCheckBox()
                bootup_test.setChecked(row_data[3])
                bootup_test.setEnabled(False)
                bootup_test.setStyleSheet("""
                    QCheckBox::indicator {
                        width: 15px;
                        height: 15px;
                    }
                    QCheckBox::indicator::unchecked {
                        background-color: white;
                        border: 1px solid white;
                    }
                    QCheckBox::indicator:checked {
                        background-color: green;
                    }
                """)
                self.center_checkbox_in_cell(row_index, 4, bootup_test)


                led_glow_test = QCheckBox()
                led_glow_test.setChecked(row_data[4])
                led_glow_test.stateChanged.connect(lambda state, idx=row_index + start_row: self.handle_led_glow_test_change(idx, state))
                led_glow_test.setStyleSheet("""
                    QCheckBox::indicator {
                        width: 15px;
                        height: 15px;
                    }
                    QCheckBox::indicator::unchecked {
                        background-color: white;
                        border: 1px solid white;
                    }
                    QCheckBox::indicator:checked {
                        background-color: green;
                    }
                """)
                self.center_checkbox_in_cell(row_index, 5, led_glow_test)

                #if checkbox.isChecked():
                    #selected_count += 1
        self.table_widget.setColumnWidth(0, 20)
        #self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        #self.table_widget.horizontalHeader().setStretchLastSection(True)
        #self.checkbox_count_label.setText(f"Selected: {selected_count}")

        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        for column in range(self.table_widget.columnCount()):
            header.setSectionResizeMode(column, QHeaderView.Stretch)

        self.update_page_info()

    '''def update_checkbox_count(self):
        self.selected_count = 0
        for row in range(self.table_widget.rowCount()):
            item = self.table_widget.item(row, 0)  # Assuming the checkbox is in column 5
            if item and item.checkState() == Qt.Checked:
                self.selected_count += 1
        self.checkbox_count_label.setText(f"{self.selected_count} MAC Addresses are selected")
        return self.selected_count
        #self.checkbox.setText(f"Select All ({self.selected_count} selected)")'''

    def update_checkbox_count(self):
        print("i am inside")
        self.selected_count = sum(1 for row_data in self.data if row_data[5])
    
        # Create a QFont object
        font = self.table_widget.horizontalHeader().checkbox.font()
    
        # Apply the font to the checkbox
        self.table_widget.horizontalHeader().checkbox.setFont(font)
        self.table_widget.horizontalHeader().checkbox.setText(f"All ({self.selected_count} selected)")
        total_rows = len(self.data)
        start_row = self.current_page * self.rows_per_page
        end_row = min(start_row + self.rows_per_page, total_rows)
        #print(f"length of self.data is {len(self.data)}")
        #end_row = (self.current_page + 1) * self.rows_per_page
        self.rows_in_a_page = end_row - start_row
        print(f"Inside the main window selected count is {self.selected_count} and rows_in a page is {self.rows_in_a_page}")
        if self.selected_count == self.rows_in_a_page:
            font.setBold(True)
            print(f"Flag before is {self.select_all_checkbox_triggered}")
            self.select_all_checkbox_triggered = True
            self.checkable_header.checkbox.setChecked(True)
            print(f"Flag after is {self.select_all_checkbox_triggered}")
            self.checkable_header.checkbox.setText(f"{self.selected_count} MAC Addresses are selected")
            #self.select_all_checkbox_triggered = False
        elif self.selected_count == 0:
            font.setBold(True)
            self.select_all_checkbox_triggered = True
            self.checkable_header.checkbox.setChecked(False)
            self.checkable_header.checkbox.setText(f"All")


        else:
            self.select_all_checkbox_triggered = False
            print(f"Flag before WHILE UNCHECKING is {self.select_all_checkbox_triggered}")

            self.checkable_header.checkbox.setChecked(False)
            print(f"Flag AFTER is {self.select_all_checkbox_triggered}")

            self.checkable_header.checkbox.setText(f"{self.selected_count} MAC Addresses are selected")
            font.setBold(False)
        #self.checkbox_count_label.setText(f"{self.selected_count} MAC Addresses are selected")
        return self.selected_count


    def load_previous_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_table()
            self.update_page_info()

    def load_next_page(self):
        total_pages = (len(self.data) + self.rows_per_page - 1) // self.rows_per_page
        if self.current_page < total_pages-1:
            self.current_page += 1
            self.update_table()
            self.update_page_info()

    
    '''def select_all_checkboxes(self, is_checked):
        #print(f"status = {is_checked}\n")
        for row in range(self.table_widget.rowCount()):
            checkbox = self.table_widget.cellWidget(row, 0).layout().itemAt(0).widget()
            #print(f"checkbox = {checkbox}")
            if checkbox:
                checkbox.setChecked(is_checked)'''
    '''def select_all_checkboxes(self, is_checked):
        # Only operate within the current page's range
        start_row = self.current_page * self.rows_per_page
        end_row = min(start_row + self.rows_per_page, len(self.data))

        for row in range(0, self.rows_per_page):
            cell_widget = self.table_widget.cellWidget(row, 0)
            if cell_widget:  # Ensure the cell widget exists
                checkbox = cell_widget.layout().itemAt(0).widget()
                if checkbox:  # Ensure the checkbox widget exists
                    checkbox.setChecked(is_checked)'''
    '''/*important def select_all_checkboxes(self, is_checked):
        # Only operate within the current page's range
        start_row = self.current_page * self.rows_per_page
        end_row = min(start_row + self.rows_per_page, len(self.data))

        for row in range(0, self.rows_per_page):
            cell_widget = self.table_widget.cellWidget(row, 0)
            if cell_widget:  # Ensure the cell widget exists
                checkbox = cell_widget.layout().itemAt(0).widget()
                if checkbox:  # Ensure the checkbox widget exists
                    checkbox.setChecked(is_checked)
        self.update_checkbox_count()'''
    
    def select_all_checkboxes(self, checked):
        self.select_all_checkbox_triggered = True  # Set the flag to indicate Select All is in action
        for row in range(0, self.rows_per_page):
            print(f"the state is {checked}")
            print(f"checked in select_all_chekbox")
            item = self.table_widget.cellWidget(row, 0)
            if item:
                #checkbox = item.findChild(QCheckBox)
                checkbox = item.layout().itemAt(0).widget()
                if checkbox:
                    checkbox.setCheckState(Qt.Checked if checked else Qt.Unchecked)
                    self.checkable_header.set_row_checked(row, checked)
        self.update_checkbox_count()

    def highlight_row(self, row_index):
        first_column_checkbox = self.table_widget.cellWidget(row_index, 0).layout().itemAt(0).widget()
        font = QFont()
        #if first_column_checkbox.isChecked():
            #font.setBold(True)
        #else:
        font.setBold(first_column_checkbox.isChecked())
        for col in range(self.table_widget.columnCount()):
            item = self.item(row_index, col)
            if item:
                item.setFont(font)

    def handle_checkbox_change(self, row_index, state):
        self.data[row_index][5] = state == Qt.Checked
        if (state == Qt.Checked):
            self.setRowColor(row_index, QColor(100, 150, 200))
        else:
            self.resetRowColor(row_index)
        print("sending to checkable_header")
        print(f"the data in main window is {self.data}")

        if not self.select_all_checkbox_triggered:
            self.checkable_header.set_row_checked(row_index, state)
        else:
            self.update_checkbox_count()
        #self.update_checkbox_count()
        #self.checkable_header.set_row_checked(row_index, state)

        #self.update_checkbox_count()

    def setRowColor(self, row, color):
        for col in range(self.table_widget.columnCount()):
            item = self.table_widget.item(row % self.rows_per_page, col)
            if not item:
                item = QTableWidgetItem()
                self.table_widget.setItem(row % self.rows_per_page, col, item)
            item.setBackground(color)

    def resetRowColor(self, row):
        if row % 2 == 0:
            color = self.palette().color(QPalette.Base)
        else:
            color = self.palette().color(QPalette.AlternateBase)

        for col in range(self.table_widget.columnCount()):
            item = self.table_widget.item(row % self.rows_per_page, col)
            if item:
                item.setBackground(color)


    '''def handle_checkbox_change(self, row_index, state):
        self.data[row_index][5] = state == Qt.Checked
        #print(f"Checkbox for row {row_index} changed to {'checked' if state ==Qt.Checked else 'unchecked'}")
        font = QFont()
        font.setBold(state == Qt.Checked)
        font.setPointSize(12)
        for col in range(self.table_widget.columnCount()):
            item = self.table_widget.item(row_index, col)
            if item:
                item.setFont(font)'''

    def handle_led_glow_test_change(self, row_index, state):
        self.data[row_index][4] = (state == Qt.Checked)


    def center_checkbox_in_cell(self, row, column, checkbox):
        cell_widget = QWidget()
        layout = QHBoxLayout(cell_widget)
        layout.addWidget(checkbox)
        layout.setAlignment(Qt.AlignCenter)
        layout.setContentsMargins(0, 0, 0, 0)
        cell_widget.setLayout(layout)
        self.table_widget.setCellWidget(row, column, cell_widget)
        if column in [0,5]:
            cell_widget.mousePressEvent = lambda event, chk=checkbox: self.cell_click_handler(chk)

    def cell_click_handler(self, checkbox):
        checkbox.toggle()    

    def handle_cell_click(self, row, col):
        # Handle clicks for the 1st and 5th columns
        if col in [0, 5]:
            cell_widget = self.table.cellWidget(row, col)
            if cell_widget:
                checkbox = cell_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.toggle()

    def download_excel(self):
        #options = QFileDialog.Options()
        #options |= QFileDialog.DontUseNativeDialog
        #file_name, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        #home_path = os.getenv('HOME')
        #print(f"home_path = {home_path}")
        path = gui_config.Download_path
        timestamp = time.strftime('%Y%m%d_%H%M%S')
        file_name = os.path.join(path, f"_output_{timestamp}.xlsx")
        #file_name = home_path + f"/Downloads/_output_{timestamp}.xlsx"
        print(f"file_name={file_name}")
        if file_name:
            self.status_label.setText('Excel sheet is downloading. Please wait...')
            self.save_data(file_name)
            self.status_label.setText(f'Excel sheet is downloaded in {file_name}.')
    
    
    def save_data(self, file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Updated Data"
        headers = ['MAC Address', 'Battery_Status', 'Switch_Press', 'Bootup_Test', 'LED Glow', 'updated_on']
        sheet.append(headers)
 
        # Set fixed column width
        column_widths = [5, 15, 15, 15, 15, 20]
        for i, column_width in enumerate(column_widths, 1):
            sheet.column_dimensions[chr(64 + i)].width = column_width
 
        # Print and add data
        current_datetime = datetime.now().strftime("%m/%d/%Y %H:%M")
        print(f"the self.data is {self.data}")
        for row_data in self.data:
            row = [
                row_data[0],
                "✔" if row_data[1] else "✘",
                "✔" if row_data[2] else "✘",
                "✔" if row_data[3] else "✘",
                "✔" if row_data[4] else "✘",
                current_datetime
            ]
            #print(row)
            sheet.append(row)
 
        # Define light green and red colors
        light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light Green
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light Red
 
        # Set cell background color based on content
        for row in sheet.iter_rows(min_row=2, max_col=5, max_row=len(self.data) + 1):
            for cell in row:
                if cell.value == "✔":
                    cell.fill = light_green_fill
                elif cell.value == "✘":
                    cell.fill = light_red_fill
 
        workbook.save(filename=file_name)
 
 
    def closeEvent(self, event):
        self.mqtt_handler.client.loop_stop()
        self.mqtt_handler.client.disconnect()
        super().closeEvent(event)


if __name__ == '__main__':
    app = QApplication(sys.argv)

    window = DarkWindow()
    window.show()
    QApplication.processEvents()


    sys.exit(app.exec_())