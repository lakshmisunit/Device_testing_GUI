#import message_capturer
import sys, os, re
import json
import paho.mqtt.client as mqtt
import ssl
import threading
from openpyxl import load_workbook, Workbook
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QProgressBar, QWidget, QLabel, QPushButton, QAction, QMenu, QComboBox, QToolBar,
    QTableWidget, QSpacerItem, QTableWidgetItem, QDialog, QFrame, QFileDialog, QCheckBox, QSizePolicy, QHeaderView, QHBoxLayout, QSpinBox, QMessageBox, QDialogButtonBox
)
from PyQt5.QtGui import QPalette, QColor, QFont, QBrush, QIcon
from PyQt5.QtCore import Qt, pyqtSignal, QThread, QSize, QEvent, QObject, QRect
from openpyxl.styles import PatternFill
from datetime import datetime
from PyQt5.QtWidgets import QMessageBox, QInputDialog
import signal, time
import gui_config
from difflib import get_close_matches, SequenceMatcher

MAC_address_key = 'Mac Address'

def signal_handler(sig, frame):
    signal_names = {signal.SIGINT: "SIGINT (Ctrl+C)",
                    signal.SIGTERM: "SIGTERM (kill)"}
    print(f"{signal_names.get(sig, f'Signal {sig}')}, but ignoring to continue execution.")

# Register signal handlers
for sig in [signal.SIGINT, signal.SIGTERM]:
    signal.signal(sig, signal_handler)

class CheckableHeader(QHeaderView):
    def __init__(self, table_widget, parent=None):
        super().__init__(Qt.Horizontal, parent)
        self.table_widget = table_widget
        self.checkbox = QCheckBox(self)
        self.checkbox.setText("All")
        self.checkbox.setStyleSheet("margin-left: 2px; text-align: center;")
        self.checkbox.stateChanged.connect(self.on_state_changed)
        self.checked_rows = set()
        self.select_all_checkbox_triggered = False

    def resizeEvent(self, event):
        super().resizeEvent(event)
        section_width = self.sectionSize(0)
        header_height = self.height()
        checkbox_width = 150
        checkbox_height = 30
        x = int((section_width - checkbox_width) / 2)
        y = int((header_height - checkbox_height) / 2)

        self.checkbox.setGeometry(QRect(x, y, checkbox_width, checkbox_height))

    def on_state_changed(self, state):
        self.select_all_checkbox_triggered = True
        is_checked = (state == Qt.Checked)
        print("checked")
        table_window = self.parent().window()
        print(f"table_window is {table_window}\n")
        if table_window:
            if(table_window.select_all_checkbox_triggered):
                table_window.select_all_states[table_window.current_page] = is_checked
                table_window.select_all_checkboxes(is_checked)
            print(state)
        if(self.select_all_checkbox_triggered) and is_checked is True:
            table_window.select_all_states[table_window.current_page] = is_checked
            table_window.select_all_checkboxes(is_checked)
        self.select_all_checkbox_triggered = False

    def update_checkbox_count(self):
        print("enetered updating for header checkbox")
        table_window = self.parent().window()
        total_rows = len(table_window.data)
        start_row = table_window.current_page * table_window.rows_per_page
        print(start_row)
        end_row = min(start_row + table_window.rows_per_page, total_rows)
        page_data = table_window.data[start_row:end_row]
        print(page_data)
        self.selected_count = sum(1 for row_data in page_data if row_data[5] and self.is_row_on_current_page(row_data))
        print(f"count of selection is : {self.selected_count}")
        font = self.checkbox.font()
        font.setBold(self.selected_count > 0)
        self.checkbox.setFont(font)
        print(end_row)
        self.rows_in_a_page = end_row - start_row
        print(self.rows_in_a_page)
        print(f"current no of rows in page is {table_window.rows_in_a_page}")
        if self.selected_count != self.rows_in_a_page:
            #self.select_all_checkbox_triggered = False
            print(f"Selected count is {self.selected_count} and rows_in_page is {table_window.rows_in_a_page}")
            self.checkbox.setText(f"All ({self.selected_count} selected)")
            self.checkbox.setChecked(False)
        elif self.selected_count == self.rows_in_a_page:
            #self.select_all_checkbox_triggered = True
            print(f"Selected count is {self.selected_count} and rows_in_page is {table_window.rows_in_a_page}")
            self.checkbox.setChecked(True)
            self.checkbox.setText(f"All ({self.selected_count} selected)")
        return self.selected_count
    
    def set_row_checked(self, row, checked):
        if checked:
            print("added")
            self.checked_rows.add(row)
        else:
            print("discarded")
            self.checked_rows.discard(row)
        self.update_checkbox_count()

    def is_row_on_current_page(self, row_data):
        """ Helper method to determine if a row belongs to the current page based on its index. """
        table_window = self.parent().window()
        row_index = table_window.data.index(row_data)
        start_row = table_window.current_page * table_window.rows_per_page
        end_row = min(start_row + table_window.rows_per_page, len(table_window.data))
        return start_row <= row_index < end_row

class TopicSelectionDialog(QDialog):
    def __init__(self, parent=None):
        super(TopicSelectionDialog, self).__init__(parent)
        self.setWindowTitle(" ")
        self.setFixedSize(600, 300)
        self.setStyleSheet("background-color: lightgrey;")
 
        # Custom Title Bar
        title_bar = QWidget(self)
        title_layout = QHBoxLayout()
        title_label = QLabel("Settings", self)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        title_layout.addWidget(title_label)
 
        title_layout.setContentsMargins(0, 5, 0, 10)  # Reduced top margin
        title_bar.setLayout(title_layout)
 
        self.layout = QVBoxLayout()
        self.layout.addWidget(title_bar)
 
        # Page Parameters
        self.rows_layout = QHBoxLayout()
        self.rows_label = QLabel("Page Size:", self)
        self.rows_label.setStyleSheet("font-size: 18px;")
        self.rows_layout.addWidget(self.rows_label)
 
        self.rows_spin_box = QSpinBox(self)
        self.rows_spin_box.setRange(1, 100)
        self.rows_spin_box.setValue(self.parent().window().rows_per_page)  # Default value
        self.rows_spin_box.setMinimumWidth(120)
        self.rows_spin_box.setAlignment(Qt.AlignCenter)  # Center-align the text
        self.rows_spin_box.setStyleSheet("height: 30px; border: 2px solid black; font-size: 18px;")
        self.rows_layout.addWidget(self.rows_spin_box)
        self.layout.addLayout(self.rows_layout)
 
        # Divider Line
        self.divider_line = QFrame(self)
        self.divider_line.setFrameShape(QFrame.HLine)
        self.divider_line.setFrameShadow(QFrame.Sunken)
        self.divider_line.setStyleSheet("color: grey;")
        self.divider_line.setFixedHeight(2)
        self.layout.addWidget(self.divider_line)
 
        # LED Glow Test Settings Title
        self.test_settings_label = QLabel("LED Glow Test Settings", self)
        self.test_settings_label.setAlignment(Qt.AlignCenter)
        self.test_settings_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        self.layout.addWidget(self.test_settings_label)
 
        # Run Test Parameters
        self.topic_layout = QHBoxLayout()
        self.topic_label = QLabel("Topic Name:", self)
        self.topic_label.setStyleSheet("font-size: 18px;")
        self.topic_layout.addWidget(self.topic_label)
 
        self.topic_combo = QComboBox()
        self.topic_combo.setEditable(True)
        self.topic_combo.lineEdit().setAlignment(Qt.AlignCenter)
        self.topic_combo.addItems(gui_config.Topic_Names)  # Example topics
        self.topic_combo.setCurrentText(self.parent().window().selected_topic)
        self.topic_combo.setMinimumWidth(200)  # Increased width
        #self.topic_combo.setAlignment(Qt.AlignCenter)
        self.topic_combo.setStyleSheet("height: 30px; border: 2px solid black; font-size: 18px; text-align: center;")
        self.topic_layout.addWidget(self.topic_combo)
        self.layout.addLayout(self.topic_layout)
 
        # Blink Duration
        self.blink_duration_layout = QHBoxLayout()
        self.blink_duration_label = QLabel("LED Blink Duration (seconds):", self)
        self.blink_duration_label.setStyleSheet("font-size: 18px;")
        self.blink_duration_layout.addWidget(self.blink_duration_label)
 
        self.blink_duration_spin_box = QSpinBox(self)
        self.blink_duration_spin_box.setRange(1, 60)
        self.blink_duration_spin_box.setValue(self.parent().window().duration)  # Default value
        self.blink_duration_spin_box.setMinimumWidth(120)
        self.blink_duration_spin_box.setAlignment(Qt.AlignCenter)
        self.blink_duration_spin_box.setStyleSheet("height: 30px; border: 2px solid black; font-size: 18px;")
        self.blink_duration_layout.addWidget(self.blink_duration_spin_box)
        self.layout.addLayout(self.blink_duration_layout)
 
        # Blink Count
        self.blink_count_layout = QHBoxLayout()
        self.blink_count_label = QLabel("LED Blink Count:", self)
        self.blink_count_label.setStyleSheet("font-size: 18px;")
        self.blink_count_layout.addWidget(self.blink_count_label)
 
        self.blink_count_spin_box = QSpinBox(self)
        self.blink_count_spin_box.setRange(1, 100)
        self.blink_count_spin_box.setValue(self.parent().window().count)  # Default value
        self.blink_count_spin_box.setMinimumWidth(120)
        self.blink_count_spin_box.setAlignment(Qt.AlignCenter)
        self.blink_count_spin_box.setStyleSheet("height: 30px; border: 2px solid black; font-size: 18px;")
        self.blink_count_layout.addWidget(self.blink_count_spin_box)
        self.layout.addLayout(self.blink_count_layout)
 
        # Adding spacing between labels and spin boxes
        for layout in [self.rows_layout, self.topic_layout, self.blink_duration_layout, self.blink_count_layout]:
            layout.setSpacing(30)  # Increased space between labels and spin boxes
 
        # Button Box
        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.button_box.setCenterButtons(True)  # Center the buttons
        self.layout.addWidget(self.button_box)
 
        self.setLayout(self.layout)
        self.installEventFilter(self)
 
    def get_selected_topic(self):
        return self.topic_combo.currentText()
 
    def get_rows_per_page(self):
        return self.rows_spin_box.value()
   
    def get_blink_duration(self):
        return self.blink_duration_spin_box.value()
 
    def get_blink_count(self):
        return self.blink_count_spin_box.value()
 
    def eventFilter(self, obj, event):
        if event.type() == QEvent.Close:
            reply = QMessageBox.question(self, "Close",
                "The changes you made may get lost. Are you sure you want to exit?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                return super().eventFilter(obj, event)
            else:
                event.ignore()
                return True
        return super().eventFilter(obj, event)



class ErrorDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Error')
        self.message_label = QLabel()
        layout = QVBoxLayout()
        layout.addWidget(self.message_label)
        self.setLayout(layout)


    def set_message(self, message):
        self.message_label.setText(message)


class MqttHandler(QThread):
    mqtt_message_signal = pyqtSignal(str, str)


    def __init__(self, broker, port, username, password):
        super().__init__()
        self.broker = broker
        self.port = port
        self.username = username
        self.password = password
        self.client = mqtt.Client(client_id="unique_client_id")
        self.client.username_pw_set(self.username, self.password)
        self.client.tls_set(cert_reqs=ssl.CERT_NONE)
        self.client.tls_insecure_set(True)
        self.client.on_connect = self.on_connect
        self.client.on_message = self.on_message


    def run(self):
        self.client.connect(self.broker, self.port, 60)
        self.client.loop_forever()


    def on_connect(self, client, userdata, flags, rc):
        if rc == 0:
            print("Connected to MQTT Broker!")
            client.subscribe("#", qos=1)
        else:
            print(f"Failed to connect, return code {rc}")

    def on_message(self, client, userdata, msg):
        topic = msg.topic
        try:
            payload = msg.payload.decode('utf-8') # Try decoding with UTF-8
        except UnicodeDecodeError:
            # Handle non-UTF-8 payloads or binary data
            payload = f"Binary data received: {msg.payload}"
    
        self.mqtt_message_signal.emit(topic, payload)

    def publish_message(self, mac_address, payload):
        topic = "LED_GLOW"
        message = {
            "devID": mac_address.upper(),
            "data": payload
        }
        self.client.publish(topic, json.dumps(message), qos=1)

class TestWorker(QThread):
    update_message = pyqtSignal(str)
    update_progress = pyqtSignal(int)
    finished = pyqtSignal()

    def __init__(self, parent, selected_rows, data, duration, count, selected_topic, mqtt_handler):
        super().__init__(parent)
        self.selected_rows = selected_rows
        self.data = data
        self.duration = duration
        self.count = count
        self.selected_topic = selected_topic
        self.mqtt_handler = mqtt_handler
        self.total_count = len(selected_rows)
        self.completed_count = 0

    def run(self):
        self.selected_count = sum(1 for row_data in self.data if row_data[5])
        message = (
            f"<div style='text-align: center;'>"
            f"<b>Publishing LED glow test, '{self.selected_topic}' for the selected {self.selected_count} devices.</b><br>"
            f"<b>Blink duration: {self.duration} seconds, Blink count: {self.total_count}</b><br>"
            "Note: Go to settings to change any of the parameters."
            "</div>"
        )
        self.update_message.emit(message)

        for row_index in self.selected_rows:
            mac_address = self.data[row_index][0]
            payload = f"LG {self.duration}, {self.count}"
            self.update_message.emit(f"Test running for {mac_address}")
            self.mqtt_handler.publish_message(mac_address, payload)
            self.completed_count += 1
            progress = int((self.completed_count / self.total_count) * 100)
            self.update_progress.emit(progress)
            time.sleep(1)
        self.finished.emit()

    '''def publish_message(self, mac_address, payload):
        topic = self.selected_topic
        message = {
            "devID": mac_address.upper(),
            "data": payload
        }
        self.mqtt_handler.client.publish(topic, json.dumps(message), qos=1)'''

class DarkWindow(QMainWindow):
    def __init__(self):
        super().__init__()


        self.setWindowTitle('Device Tester')
        self.setGeometry(100, 100, 1200, 800)


        palette = QPalette()
        base_color = QColor(35, 35, 35)
        alternate_color = QColor(53, 53, 53)
        palette.setColor(QPalette.Window, QColor(53, 53, 53))
        palette.setColor(QPalette.WindowText, QColor(255, 255, 255))
        palette.setColor(QPalette.Base, QColor(35, 35, 35))
        palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
        palette.setColor(QPalette.ToolTipBase, QColor(255, 255, 255))
        palette.setColor(QPalette.ToolTipText, QColor(255, 255, 255))
        palette.setColor(QPalette.Text, QColor(255, 255, 255))
        palette.setColor(QPalette.Button, QColor(53, 53, 53))
        palette.setColor(QPalette.ButtonText, QColor(255, 255, 255))
        palette.setColor(QPalette.BrightText, QColor(255, 0, 0))
        palette.setColor(QPalette.Link, QColor(42, 130, 218))
        palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        palette.setColor(QPalette.HighlightedText, QColor(0, 0, 0))


        self.setPalette(palette)

        self.setMinimumSize(400, 200)
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)
        self.current_page = 0
        self.rows_per_page = gui_config.rows_in_page
        self.total_rows = 0
        self.select_all_states = {}
        self.prev_pages = 1
        self.rows_in_a_page = 0
        self.duration = gui_config.blink_duration
        self.count = gui_config.blink_count

        self.select_all_checkbox_triggered = False
        self.layout = QVBoxLayout(central_widget)

        self.upload_done = False

        self.file_label = QLabel('')
        self.file_label.setStyleSheet("color: white;")
        self.layout.addWidget(self.file_label)


        self.table_widget = QTableWidget(self)
        self.table_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.layout.addWidget(self.table_widget)
        self.table_widget.setVisible(False)
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setStyleSheet("""
            QTableWidget {
                alternate-background-color: #444444;
                background-color: #2B2B2B;
            }
            QHeaderView: {
                background-color: #353535;
                color: white;
            }
        """)

        self.checkable_header = CheckableHeader(self.table_widget)  # Keep a reference
        self.table_widget.setHorizontalHeader(self.checkable_header)

        self.status_label = QLabel()
        self.status_label.setStyleSheet("color: white;")
        self.layout.addWidget(self.status_label)
        
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setVisible(False)
        self.layout.addWidget(self.progress_bar)

        central_widget.setLayout(self.layout)
        self.toolbar = QToolBar()
        self.toolbar.setStyleSheet("""
            QToolBar {
                border: 2px solid white; /* Set the same background color for the toolbar */
                padding: 0px;
                spacing: 3px;
                height: 70px;
            }
            QToolButton {
                background-color: transparent;  /* Ensure buttons have the same background color */
                color: white;
                font-size: 12px;         /* Increase the font size */
                margin: 0px;
                padding: 2px;            /* Add some padding to avoid cramped text */
            }
            QToolButton::icon {
            padding-top: 5px;
            }
        """)
        self.toolbar.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)
        self.toolbar.setMovable(False)
        self.toolbar.setFloatable(False)
        self.addToolBar(Qt.TopToolBarArea, self.toolbar)
        self.toolbar.setIconSize(QSize(48,48))
        self.upload_action = QAction(QIcon('upload3.png'), 'Upload', self)
        self.upload_action.triggered.connect(self.upload_file)
        self.run_test_action = QAction(QIcon('run.png'), 'Run Test', self)
        #self.run_test_action.triggered.connect(self.run_test)
        self.run_test_action.triggered.connect(self.run_test_action_triggered)
        self.download_action = QAction(QIcon('download.png'), 'Download', self)
        self.download_action.triggered.connect(self.download_excel)
        self.settings_action = QAction(QIcon('images.png'), 'Settings', self)
        self.settings_action.triggered.connect(self.open_settings_dialog)
        self.toolbar.addAction(self.upload_action)
        self.toolbar.addAction(self.run_test_action)
        self.toolbar.addAction(self.download_action)
        self.toolbar.addAction(self.settings_action)

        self.nav_layout = QHBoxLayout()
        self.nav_layout.setContentsMargins(0, 0, 0, 0)  # No margins around the layout
        self.nav_layout.setSpacing(10)  # No spacing between items

        # Add stretchable space before the previous button
        self.nav_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))

        self.page_info_label = QLabel(self)
        self.page_info_label.setAlignment(Qt.AlignCenter)
        self.page_info_label.setStyleSheet("color: white; font-size: 14px;")  # Padding for spacing

        self.first_button = QPushButton(self)
        self.first_button.setIcon(QIcon('first_page.png'))
        self.first_button.clicked.connect(self.go_to_first_page)
        self.first_button.setVisible(False)
        self.first_button.setStyleSheet("border: none; background: none;")

        self.prev_button = QPushButton(self)
        self.prev_button.setIcon(QIcon('left.png'))
        self.prev_button.clicked.connect(self.load_previous_page)
        self.prev_button.setVisible(False)
        self.prev_button.setStyleSheet("border: none; background: none;")  # Transparent background


        self.next_button = QPushButton(self)
        self.next_button.setIcon(QIcon('right.png'))        
        self.next_button.clicked.connect(self.load_next_page)
        self.next_button.setVisible(False)
        self.next_button.setStyleSheet("border: none; background: none;")  # Transparent background

        self.last_button = QPushButton(self)
        self.last_button.setIcon(QIcon('last_page.png'))
        self.last_button.clicked.connect(self.go_to_last_page)
        self.last_button.setVisible(False)
        self.last_button.setStyleSheet("border: none; background: none;")

        self.first_button.setFixedSize(30, 30)
        self.prev_button.setFixedSize(30, 30)  # Set a fixed size for the left icon
        self.next_button.setFixedSize(30, 30)
        self.last_button.setFixedSize(30, 30)

        self.nav_layout.addWidget(self.first_button, alignment=Qt.AlignCenter)
        self.nav_layout.addWidget(self.prev_button, alignment=Qt.AlignCenter)
        self.nav_layout.addWidget(self.page_info_label, alignment=Qt.AlignCenter)
        self.nav_layout.addWidget(self.next_button, alignment=Qt.AlignCenter)
        self.nav_layout.addWidget(self.last_button, alignment=Qt.AlignCenter)
        
        self.nav_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
        self.layout.addLayout(self.nav_layout)

        self.selected_topic = 'LED_GLOW'

        self.installEventFilter(self) 
        self.data = []

        self.MQTT_BROKER = "103.162.246.109"
        #self.MQTT_BROKER = "192.168.100.18"
        self.MQTT_PORT = 8883
        self.MQTT_USERNAME = "mqtt"
        self.MQTT_PASSWORD = "mqtt2022"

        self.mqtt_handler = MqttHandler(self.MQTT_BROKER, self.MQTT_PORT, self.MQTT_USERNAME, self.MQTT_PASSWORD)
        #self.boot_up_logic()
        self.mqtt_handler.mqtt_message_signal.connect(self.handle_mqtt_message)
        self.mqtt_handler.start()


        self.selected_count = 0
        print(f"selected devices count = {self.selected_topic}")
        self.topic_dialog = TopicSelectionDialog(self)

    def run_test_action_triggered(self):
        selected_rows = self.get_selected_rows()
        self.selected_count = sum(1 for row_data in self.data if row_data[5])
        print(f"selected_count is {self.selected_count}")
        if not selected_rows:
            error_dialog = QMessageBox(self)
            error_dialog.setIcon(QMessageBox.Warning)
            error_dialog.setText("Please select at least one row.")
            error_dialog.setWindowTitle("Error")
            error_dialog.exec_()
            return 
    
        message = (
        f"<div style='text-align: center;'>"
        f"<b>Publishing LED glow test, '{self.selected_topic}' for the selected {self.selected_count} devices.</b><br>"
        f"<b>Blink duration: {self.duration} seconds, Blink count: {self.count}</b><br>"
        "Note: Go to settings to change any of the parameters."
        "</div>"
    )
        custom_dialog = QDialog(self)
        custom_dialog.setWindowTitle("Run Test")
        custom_dialog.setFixedSize(550, 150)

        layout = QVBoxLayout()
        message_label = QLabel(message)
        message_label.setWordWrap(True)
        #message_label.setStyleSheet()
        layout.addWidget(message_label)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, custom_dialog)
        layout.addWidget(button_box)
        custom_dialog.setLayout(layout)

        button_box.accepted.connect(custom_dialog.accept)
        button_box.rejected.connect(custom_dialog.reject)

        reply = custom_dialog.exec_()

        if reply == QDialog.Accepted:
            self.test_worker = TestWorker(
                self,
                selected_rows = selected_rows,
                data = self.data,
                duration = self.duration,
                count = self.count,
                selected_topic = self.selected_topic, 
                mqtt_handler = self.mqtt_handler
                )
            self.test_worker.update_message.connect(self.on_test_update_message)
            self.progress_bar.setVisible(True)
            self.test_worker.update_progress.connect(self.update_progress_bar)
            self.test_worker.finished.connect(self.on_test_finished)
            self.test_worker.start()
    
    def update_progress_bar(self, progress):
        self.progress_bar.setValue(progress)

    def on_test_update_message(self, message):
        self.status_label.setText(message)
    
    def on_test_finished(self):
        self.status_label.setText("Test Finished")
        self.progress_bar.setValue(100)

    def get_selected_rows(self):
        selected_rows = [index for index, row in enumerate(self.data) if row[5]]
        return selected_rows
    
    def update_page_info(self):
        total_pages = (self.total_rows + self.rows_per_page - 1) // self.rows_per_page
        if total_pages > 1:
            self.first_button.setVisible(True)
            self.prev_button.setVisible(True)
            self.next_button.setVisible(True)
            self.last_button.setVisible(True)
            print("Added first and last button")
            self.page_info_label.setText(f"Page {self.current_page + 1} of {total_pages}")
            self.first_button.setEnabled(self.current_page > 0)
            self.prev_button.setEnabled(self.current_page > 0)
            self.next_button.setEnabled(self.current_page < total_pages - 1)
            self.last_button.setEnabled(self.current_page < total_pages - 1)
        else:
            self.page_info_label.setText('')
            self.first_button.setVisible(False)
            self.prev_button.setVisible(False)
            self.next_button.setVisible(False)
            self.last_button.setVisible(False)
        
        # Optional: Set a different style for disabled buttons
        if not self.first_button.isEnabled():
            self.first_button.setStyleSheet("opacity: 0.5;")
        else:
            self.first_button.setStyleSheet("")

        if not self.prev_button.isEnabled():
            self.prev_button.setStyleSheet("opacity: 0.5;")  # Make it look disabled
        else:
            self.prev_button.setStyleSheet("")  # Reset style

        if not self.next_button.isEnabled():
            self.next_button.setStyleSheet("opacity: 0.5;")  # Make it look disabled
        else:
            self.next_button.setStyleSheet("")  # Reset style

        if not self.last_button.isEnabled():
            self.last_button.setStyleSheet("opacity: 0.5;")
        else:
            self.last_button.setStyleSheet("")
        

    def go_to_first_page(self):
        self.current_page = 0
        self.update_table()
        self.update_page_info()
        self.update_highlights()
        #self.load_page_data()

    def go_to_last_page(self):
        total_page = (self.total_rows + self.rows_per_page - 1) // self.rows_per_page
        self.current_page = total_page - 1
        self.update_table()
        self.update_page_info()
        self.update_highlights()
        #self.load_page_data()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Close:
            reply = QMessageBox.question(self, "Exit", "Exit GUI?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                return super().eventFilter(obj, event)
            else:
                event.ignore()
                return True
        return super().eventFilter(obj, event)
    
    def handle_mqtt_message(self, topic, payload):
        try:
            print(f"Topic = {topic}")
            print(f"receieved payload is :{payload}")
            message = json.loads(payload)
            dev_id = message.get('devID')
            print(f"dev_id is {dev_id}")
            if dev_id and self.upload_done:
                #mac_address = ':'.join([dev_id[i:i + 2] for i in range(0, len(dev_id), 2)])
                mac_address = dev_id
                self.update_table_with_mqtt_data(mac_address, message.get('data'))
            else:
                print("Received message, but no devID found in payload")

            #print(f"mac_address = {mac_address}, data = {message.get('data')}")
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON: {e}")
        #QApplication.processEvents()
        

    def open_settings_dialog(self):
        dialog = TopicSelectionDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            self.selected_topic = dialog.get_selected_topic()
            self.rows_per_page = dialog.get_rows_per_page()
            self.duration = dialog.get_blink_duration()
            self.count = dialog.get_blink_count()
            if(self.upload_done):
                self.status_label.setText(f"Selected Topic for LED Test: {self.selected_topic}")
                self.update_page_info()
                self.update_table()
                print(f"the total rows are {self.total_rows}")
                total_pages = (self.total_rows + self.rows_per_page - 1) // self.rows_per_page
                print(f"Total_pages before are {self.prev_pages} and now it is {total_pages}") 
                if(total_pages < self.prev_pages):
                    self.go_to_last_page() 
                if(total_pages == 0):
                    QMessageBox.warning(self, "No Valid Addresses", "No valid addresses are found in the given Excel sheet.")
                self.prev_pages = total_pages
                    
    def run_test(self):
        selected_rows = self.get_selected_rows()
        if not selected_rows:
            error_dialog = ErrorDialog(self)
            error_dialog.set_message("Please select at least one row.")
            error_dialog.exec_()
            return
        
        self.selected_count = sum(1 for row_data in self.data if row_data[5])

        message = (
        f"<div style='text-align: center;'>"
        f"<b>Publishing LED glow test, '{self.selected_topic}' for the selected {self.selected_count} devices.</b><br>"
        f"<b>Blink duration: {self.duration} seconds, Blink count: {self.count}</b><br>"
        "Note: Go to settings to change any of the parameters."
        "</div>"
    )
 
        # Create a custom dialog for confirmation
        custom_dialog = QDialog(self)
        custom_dialog.setWindowTitle("Run Test")
        custom_dialog.setFixedSize(550, 150)  # Adjust size as necessary
 
        layout = QVBoxLayout()
   
        message_label = QLabel(message)
        message_label.setWordWrap(True)
        message_label.setStyleSheet("font-size: 15px")
        layout.addWidget(message_label)
 
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, custom_dialog)
        button_box.setCenterButtons(True)
        layout.addWidget(button_box)
 
        custom_dialog.setLayout(layout)
 
        # Connect buttons
        button_box.accepted.connect(custom_dialog.accept)
        button_box.rejected.connect(custom_dialog.reject)
 
        # Show the custom dialog
        reply = custom_dialog.exec_()
 
        print(f"selected devices count = {self.selected_count}")

        if reply == QDialog.Accepted:
            for row_index in selected_rows:
                mac_address = self.data[row_index][0]
                mac_address_without_colon = ''.join(mac_address.split(':'))
                payload = f"LG {self.duration},{self.count}"
                self.publish_message(mac_address, payload)
            QApplication.processEvents()
        else:
            confirmation_reply = QMessageBox.question(self, "Confirm Exit", "Are you sure you want to quit running the test?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            #return  # User chose to cancel

            if confirmation_reply == QMessageBox.Yes:
                return 

        

    '''def get_selected_rows(self):
        selected_rows = []
        for row_index, row_data in enumerate(self.data):
            if row_data[5]:
                #print(row_data[5])
                selected_rows.append(row_index)
        return selected_rows'''
    
    def publish_message(self, mac_address, payload):
        topic = self.selected_topic
        message = {
            "devID": mac_address.upper(),
            "data": payload
        }
        self.mqtt_handler.client.publish(topic, json.dumps(message), qos=1)
        QApplication.processEvents()
        #print(f"Published message '{message}' to topic '{topic}'")
 
        
    def update_table_with_mqtt_data(self, mac_address, data):
        data_found = False
        mac_address_lower = mac_address.lower()
    
        # Check if 'LG' appears in the data
        #is_LG_data = any(part.strip().lower().startswith("lg") for part in data.split(','))
        #print(f"LG data is {is_LG_data}")

        for row_index, row_data in enumerate(self.data):
            if row_data[0] is not None and row_data[0].lower() == mac_address_lower:
                data_found = True
                self.update_checkbox(row_index, data)
                print(data)

                # Only move the row if the data does not contain "LG"
                if any(row_data[1:]):
                    print("entering to pop")
                    row = self.data.pop(row_index)
                    self.data.insert(0, row)
                break

        if not data_found:
            print(data)
            new_row = [mac_address, False, False, False, False, False]
        
            print(f"before popping no of rows will be {len(self.data)}")
            last_row = self.data.pop(len(self.data)-1)
            print(f"After popping the last row, no of rows will be {len(self.data)}")
            print(f"Last row is {last_row}")
            print("pushed to first row")
            self.data.insert(0, new_row)  # Push new row to the top
            print(f"before adding the popped row to last, no of rows will be {len(self.data)}")
            print("adding row to last")
            self.data.append(last_row)
            self.total_rows = len(self.data)
            print(f"after adding the popped row to last no of rows will be {len(self.data)}")
                #self.update_page_info()
            self.update_page_info()
            self.update_checkbox(0, data)
            # Add the new row to the end if data is LG
            '''if is_LG_data:
                self.data.append(new_row)  # Append to the end for LG data
            else:
                print(f"before popping no of rows will be {len(self.data)}")
                last_row = self.data.pop(len(self.data)-1)
                print(f"After popping the last row, no of rows will be {len(self.data)}")
                print(f"Last row is {last_row}")
                print("pushed to first row")
                self.data.insert(0, new_row)  # Push new row to the top
                print(f"before adding the popped row to last, no of rows will be {len(self.data)}")
                print("adding row to last")
                self.data.append(last_row)
                self.total_rows = len(self.data)
                print(f"after adding the popped row to last no of rows will be {len(self.data)}")
            self.update_page_info()
            self.update_checkbox(len(self.data)-1 if is_LG_data else 0, data)'''

        self.update_table()



    def update_checkbox(self, row_index, data):
        if data == 100:
            self.data[row_index][1] = True
            #self.data[row_index][3] = True
            move_row = True
        if data == "BootUp":
            print("Boot up is captured")
            self.data[row_index][3] = True
            move_row = True
        if data == "SW":
            self.data[row_index][2] = True
            move_row = True


    def upload_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx)", options=options)
        if file_name:
            self.load_data(file_name)
            self.status_label.setText('Excel sheet is uploading. Please wait...')
            self.file_label.setText(f'Selected File: {file_name}')
            font = self.table_widget.horizontalHeader().checkbox.font()
            font.setBold(False)
            self.table_widget.horizontalHeader().checkbox.setFont(font)
            self.table_widget.horizontalHeader().checkbox.setText(f"All")
            self.select_all_states = {}
            self.update_table()
            self.table_widget.setVisible(True)
            self.download_action.setVisible(True)
            self.run_test_action.setVisible(True)
            total_pages = (self.total_rows + self.rows_per_page - 1) // self.rows_per_page
            print(f"Previous total pages are {self.prev_pages} and current total pages are {total_pages}")
            if(total_pages == 1):
                self.go_to_first_page()
            elif(total_pages > self.prev_pages):
                self.first_button.setVisible(True)
                self.prev_button.setVisible(True)
                self.next_button.setVisible(True)
                self.last_button.setVisible(True)
            elif(total_pages == 0):
                QMessageBox.warning(self, "No Valid Addresses", "No valid addresses are found in the given Excel sheet.")
            elif(total_pages<self.prev_pages):
                self.go_to_last_page()
            else:
                self.go_to_first_page()
            
            self.status_label.setText('Upload done.')
            self.upload_done = True
            self.prev_pages = total_pages

    def find_mac_address_column(self, sheet):
        possible_headers = {'mac address', 'mac-address', 'mac_address'}
        headers = [cell.value.strip().lower() for cell in sheet[1] if cell.value]
    
        print(f"Headers found: {headers}")
 
        best_match_idx = None
        highest_ratio = 0.0
 
        for idx, header in enumerate(headers):
            print(f"Checking header: {header}")
            match = get_close_matches(header, possible_headers, n=1, cutoff=0.8)
            print(f"Match found: {match}")
        
            if match:
                ratio = SequenceMatcher(None, header, match[0]).ratio()
                print(f"Similarity ratio: {ratio}")
                if ratio > highest_ratio:
                    highest_ratio = ratio
                    best_match_idx = idx
 
        if best_match_idx is not None:
            print(f"Best match column index: {best_match_idx}")
        else:
            print("No match found.")
 
        return best_match_idx


    def load_data(self, file_name):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active

        mac_addresses = set()
        duplicates = set()
        complete_mac_address = set()
        mac_address_column = self.find_mac_address_column(sheet)
        print(mac_address_column)
        self.data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            mac_address = row[mac_address_column]
            if mac_address:
                normalized_mac = mac_address.strip().lower()
                complete_mac_address.add(normalized_mac)
                if normalized_mac in mac_addresses:
                    duplicates.add(normalized_mac)
                else:
                    mac_addresses.add(normalized_mac)

    # Print unique MAC addresses
        print("Unique MAC addresses:")
        for mac in complete_mac_address:
            print(mac)

        if duplicates:
            reply = QMessageBox.question(
                self, 'Duplicate MAC Addresses',
                f"Duplicate MAC addresses found: {', '.join(duplicates)}.\n\nDo you want to remove the duplicates?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                self.remove_duplicates_load(complete_mac_address)
            else:
                self._load_data_from_file(file_name, mac_address_column, remove_invalid=True)
        else:
            self._load_data_from_file(file_name, mac_address_column, remove_invalid=True)
        self.total_rows = len(self.data)
    
    def remove_duplicates_load(self, mac_address_all):
        self.data = []
        for mac in mac_address_all:
            if self.is_valid_mac_address(mac.strip().upper()):
                    self.data.append([mac.strip().upper(), False, False, False, False, False])
            
    def _load_data_from_file(self, file_name, mac_address_column, remove_invalid=True):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active
        self.data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            mac_address = row[mac_address_column]
            if mac_address:
                normalized_mac = mac_address.strip().lower()
                if remove_invalid:
                    if self.is_valid_mac_address(normalized_mac.upper()):
                        print(f"{normalized_mac} is valid")
                        self.data.append([normalized_mac.upper(), False, False, False, False, False])

    def is_valid_mac_address(self, mac_address):
        print(f"{mac_address} is checking for validity")
        if not re.match(r'^([0-9A-Fa-f]{2}:){5}[0-9A-Fa-f]{2}$', mac_address):
            return False
        if not mac_address.startswith("AB:05:03"):
            return False
        return True
    
    def update_table(self):
        self.table_widget.clear() # Clears all items from the table widget
        self.table_widget.setRowCount(len(self.data)) # Sets the number of rows in the table to the length of self.data
        self.table_widget.setColumnCount(6) # Sets the number of columns in the table to the length of self.data
        self.table_widget.setAlternatingRowColors(True) # Enables alternating row colors for the table
        self.table_widget.setHorizontalHeaderLabels(['', 'MAC Address', 'Battery Status', 'Switch Press', 'Bootup Test', 'LED Glow']) # Sets the horizontal header labels for the columns
        
        start_row = self.current_page * self.rows_per_page
        end_row = (self.current_page + 1) * self.rows_per_page
        page_data = self.data[start_row:end_row]
        select_all_checked = self.select_all_states.get(self.current_page, False)
        self.table_widget.setRowCount(len(page_data))

        self.checkable_header.checkbox.setChecked(select_all_checked)
        self.table_widget.horizontalHeader().setVisible(True)

        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)

        for row_index, row_data in enumerate(page_data):
            if len(row_data) >= 6:
                # Creates and configure a checkbox for the first column
                checkbox= QCheckBox()
                checkbox.setChecked(row_data[5])
                checkbox.stateChanged.connect(lambda state, idx=row_index + start_row: self.handle_checkbox_change(idx, state))
                self.center_checkbox_in_cell(row_index, 0, checkbox)

                mac_item = QTableWidgetItem(row_data[0])
                font = mac_item.font()
                font.setPointSize(12)
                mac_item.setFont(font)
                mac_item.setForeground(QBrush(QColor(255, 255, 255)))
                mac_item.setTextAlignment(Qt.AlignCenter)
                mac_item.setFlags(mac_item.flags() & ~Qt.ItemIsEditable | Qt.ItemIsSelectable)
                self.table_widget.setItem(row_index, 1, mac_item)

                # Created and configure a checkbox for the Battery Status column
                battery_status = QCheckBox()
                battery_status.setChecked(row_data[1])
                battery_status.setEnabled(False)
                battery_status.setStyleSheet("""
                    QCheckBox::indicator {
                        width: 15px;
                        height: 15px;
                    }
                    QCheckBox::indicator::unchecked {
                        background-color: white;
                        border: 1px solid white;
                    }
                    QCheckBox::indicator:checked {
                        background-color: green;
                    }
                """)
                self.center_checkbox_in_cell(row_index, 2, battery_status)


                switch_press = QCheckBox()
                switch_press.setChecked(row_data[2])
                switch_press.setEnabled(False)
                switch_press.setStyleSheet("""
                    QCheckBox::indicator {
                        width: 15px;
                        height: 15px;
                    }
                    QCheckBox::indicator::unchecked {
                        background-color: white;
                        border: 1px solid white;
                    }
                    QCheckBox::indicator:checked {
                        background-color: green;
                    }
                """)
                self.center_checkbox_in_cell(row_index, 3, switch_press)


                bootup_test = QCheckBox()
                bootup_test.setChecked(row_data[3])
                bootup_test.setEnabled(False)
                bootup_test.setStyleSheet("""
                    QCheckBox::indicator {
                        width: 15px;
                        height: 15px;
                    }
                    QCheckBox::indicator::unchecked {
                        background-color: white;
                        border: 1px solid white;
                    }
                    QCheckBox::indicator:checked {
                        background-color: green;
                    }
                """)
                self.center_checkbox_in_cell(row_index, 4, bootup_test)

                led_glow_test = QCheckBox()
                led_glow_test.setChecked(row_data[4])
                led_glow_test.stateChanged.connect(lambda state, idx=row_index + start_row: self.handle_led_glow_test_change(idx, state))
                led_glow_test.setStyleSheet("""
                    QCheckBox::indicator {
                        width: 15px;
                        height: 15px;
                    }
                    QCheckBox::indicator::unchecked {
                        background-color: white;
                        border: 1px solid white;
                    }
                    QCheckBox::indicator:checked {
                        background-color: green;
                    }
                """)
                self.center_checkbox_in_cell(row_index, 5, led_glow_test)


        self.table_widget.setColumnWidth(0, 20)

        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        for column in range(self.table_widget.columnCount()):
            header.setSectionResizeMode(column, QHeaderView.Stretch)

        self.update_page_info()
    
    def is_row_on_current_page(self, row_data):
        row_index = self.data.index(row_data)
        start_row = self.current_page * self.rows_per_page
        end_row = min(start_row + self.rows_per_page, len(self.data))
        return start_row <= row_index < end_row
    
    def update_checkbox_count(self):
        print("i am inside")
        start_row = self.current_page * self.rows_per_page
        end_row = (self.current_page + 1) * self.rows_per_page

        page_data = self.data[start_row: end_row]
        print(page_data)
        self.selected_count = sum(1 for row_data in page_data if row_data[5] and self.is_row_on_current_page(row_data))

        font = self.table_widget.horizontalHeader().checkbox.font()
    
        self.table_widget.horizontalHeader().checkbox.setFont(font)
        total_rows = len(self.data)
        start_row = self.current_page * self.rows_per_page
        end_row = min(start_row + self.rows_per_page, total_rows)
        self.rows_in_a_page = end_row - start_row
        print(f"Inside the main window selected count is {self.selected_count} and rows_in a page is {self.rows_in_a_page}")
        if self.selected_count == self.rows_in_a_page:
            font.setBold(True)
            print(f"Flag before is {self.select_all_checkbox_triggered}")
            self.select_all_checkbox_triggered = True
            self.checkable_header.checkbox.setChecked(True)
            print(f"Flag after is {self.select_all_checkbox_triggered}")
            self.checkable_header.checkbox.setText(f"{self.selected_count} MAC Addresses are selected")
        elif self.selected_count == 0:
            font.setBold(True)
            self.select_all_checkbox_triggered = True
            self.checkable_header.checkbox.setChecked(False)
            self.checkable_header.checkbox.setText(f"All")
            font.setBold(False)


        elif self.selected_count != self.rows_in_a_page:
            self.select_all_checkbox_triggered = False
            print(f"Flag before WHILE UNCHECKING is {self.select_all_checkbox_triggered}")

            self.checkable_header.checkbox.setChecked(False)
            print(f"Flag AFTER is {self.select_all_checkbox_triggered}")

            self.checkable_header.checkbox.setText(f"{self.selected_count} MAC Addresses are selected")
            font.setBold(False)
        return self.selected_count


    def update_highlights(self):
        start_row = self.current_page * self.rows_per_page
        end_row = (self.current_page + 1) * self.rows_per_page
        page_data = self.data[start_row: end_row]
        print(f"page_data is {page_data}")
        for row_data in page_data:
            if row_data[5]: #and self.is_row_on_current_page(row_data):
                self.setRowColor(page_data.index(row_data), QColor(100, 150, 200))

    def load_previous_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_table()
            self.update_checkbox_count()
            self.update_page_info()
            self.update_highlights()


    def load_next_page(self):
        total_pages = (len(self.data) + self.rows_per_page - 1) // self.rows_per_page
        if self.current_page < total_pages-1:
            self.current_page += 1
            self.update_table()
            self.update_checkbox_count()
            self.update_page_info()
            self.update_highlights()
    
    def select_all_checkboxes(self, checked):
        self.select_all_checkbox_triggered = True  # Set the flag to indicate Select All is in action
        for row in range(0, self.rows_per_page):
            print(f"the state is {checked}")
            print(f"checked in select_all_chekbox")
            item = self.table_widget.cellWidget(row, 0)
            if item:
                #checkbox = item.findChild(QCheckBox)
                checkbox = item.layout().itemAt(0).widget()
                if checkbox:
                    checkbox.setCheckState(Qt.Checked if checked else Qt.Unchecked)
                    self.checkable_header.set_row_checked(row, checked)
        self.update_checkbox_count()

    def highlight_row(self, row_index):
        first_column_checkbox = self.table_widget.cellWidget(row_index, 0).layout().itemAt(0).widget()
        font = QFont()
        font.setBold(first_column_checkbox.isChecked())
        for col in range(self.table_widget.columnCount()):
            item = self.item(row_index, col)
            if item:
                item.setFont(font)

    def handle_checkbox_change(self, row_index, state):
        self.data[row_index][5] = state == Qt.Checked
        if (state == Qt.Checked) and self.is_row_on_current_page(self.data[row_index]):
            self.setRowColor(row_index % self.rows_per_page, QColor(100, 150, 200))
        else:
            self.resetRowColor(row_index)
        if not self.select_all_checkbox_triggered:
            self.checkable_header.set_row_checked(row_index, state)
        else:
            self.update_checkbox_count()


    def setRowColor(self, row, color):
        for col in range(self.table_widget.columnCount()):
            computed_row = row%self.rows_per_page 
            print(f"computed row is {computed_row}")

            item = self.table_widget.item(row % self.rows_per_page, col)
            print(f"{col} has {item} but the column2 has this item {self.table_widget.item(2, col)}")
            #computed_row = row % self.rows_per_page
            print(f"using another variable column 2 has {self.table_widget.item(computed_row, col)}")
            if not item:
                item = QTableWidgetItem()
                self.table_widget.setItem(row % self.rows_per_page, col, item)
            item.setBackground(color)

    def resetRowColor(self, row):
        if row % 2 == 0:
            color = self.palette().color(QPalette.Base)
        else:
            color = self.palette().color(QPalette.AlternateBase)

        for col in range(self.table_widget.columnCount()):
            item = self.table_widget.item(row % self.rows_per_page, col)
            if item:
                item.setBackground(color)

    def handle_led_glow_test_change(self, row_index, state):
        self.data[row_index][4] = (state == Qt.Checked)


    def center_checkbox_in_cell(self, row, column, checkbox):
        cell_widget = QWidget()
        layout = QHBoxLayout(cell_widget)
        layout.addWidget(checkbox)
        layout.setAlignment(Qt.AlignCenter)
        layout.setContentsMargins(0, 0, 0, 0)
        cell_widget.setLayout(layout)
        self.table_widget.setCellWidget(row, column, cell_widget)
        if column in [0,5]:
            cell_widget.mousePressEvent = lambda event, chk=checkbox: self.cell_click_handler(chk)

    def cell_click_handler(self, checkbox):
        checkbox.toggle()    

    def handle_cell_click(self, row, col):
        # Handle clicks for the 1st and 5th columns
        if col in [0, 5]:
            cell_widget = self.table.cellWidget(row, col)
            if cell_widget:
                checkbox = cell_widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.toggle()

    def download_excel(self):
        path = gui_config.Download_path
        timestamp = time.strftime('%Y%m%d_%H%M%S')
        file_name = os.path.join(path, f"_output_{timestamp}.xlsx")
        #file_name = home_path + f"/Downloads/_output_{timestamp}.xlsx"
        print(f"file_name={file_name}")
        if file_name:
            self.status_label.setText('Excel sheet is downloading. Please wait...')
            self.save_data(file_name)
            self.status_label.setText(f'Excel sheet is downloaded in {file_name}.')
    
    
    def save_data(self, file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Updated Data"
        headers = ['MAC Address', 'Battery_Status', 'Switch_Press', 'Bootup_Test', 'LED Glow', 'updated_on']
        sheet.append(headers)
 
        # Set fixed column width
        column_widths = [5, 15, 15, 15, 15, 20]
        for i, column_width in enumerate(column_widths, 1):
            sheet.column_dimensions[chr(64 + i)].width = column_width
 
        # Print and add data
        current_datetime = datetime.now().strftime("%m/%d/%Y %H:%M")
        print(f"the self.data is {self.data}")
        for row_data in self.data:
            row = [
                row_data[0],
                "✔" if row_data[1] else "✘",
                "✔" if row_data[2] else "✘",
                "✔" if row_data[3] else "✘",
                "✔" if row_data[4] else "✘",
                current_datetime
            ]
            #print(row)
            sheet.append(row)
 
        # Define light green and red colors
        light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light Green
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light Red
 
        # Set cell background color based on content
        for row in sheet.iter_rows(min_row=2, max_col=5, max_row=len(self.data) + 1):
            for cell in row:
                if cell.value == "✔":
                    cell.fill = light_green_fill
                elif cell.value == "✘":
                    cell.fill = light_red_fill
 
        workbook.save(filename=file_name)
 
 
    def closeEvent(self, event):
        self.mqtt_handler.client.loop_stop()
        self.mqtt_handler.client.disconnect()
        super().closeEvent(event)


if __name__ == '__main__':
    app = QApplication(sys.argv)

    window = DarkWindow()
    window.show()
    QApplication.processEvents()


    sys.exit(app.exec_())